"""Build a clean document list for SP-01415 Snorre A SVP085 to discuss with Honeywell.

No dates. The list mirrors the actual delivered document package from the prior
prover projects (Hugin A SP-00525 / Valhall SP-00577) at file-level granularity,
plus Snorre A-specific additions (modular split, NORSOK painting, SAT, free-issue
items).

Saved to OneDrive shared project folder under:
    05 Dokumentasjon / 03.Dok sendt leverandør for oppdatering / .
"""

import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from _paths import SHARED
PROJECT = SHARED / 'Zigma360' / 'Projects' / 'SP-01415 Small Volume Prover Snorre A'
DEST = PROJECT / '05 Dokumentasjon' / '03.Dok sendt leverandør for oppdatering'
DEST.mkdir(parents=True, exist_ok=True)
OUTPUT = DEST / 'SP-01415_Document_List_for_Honeywell_Review.xlsx'

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
TITLE_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE_FONT = Font(bold=True, color='FFFFFF', size=14)

HEADER_FILL = PatternFill('solid', fgColor='2E75B6')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)

SECTION_FILL = PatternFill('solid', fgColor='DDEBF7')
SECTION_FONT = Font(bold=True, color='1F4E78', size=11)

SUBSECTION_FILL = PatternFill('solid', fgColor='F2F2F2')
SUBSECTION_FONT = Font(bold=True, color='000000', size=10)

OWNER_FILLS = {
    'HW': PatternFill('solid', fgColor='F8CBAD'),
    'SF': PatternFill('solid', fgColor='C6E0B4'),
    'HW + SF': PatternFill('solid', fgColor='FFE699'),
    'SF/Sub': PatternFill('solid', fgColor='C6E0B4'),
    'HW/Sub': PatternFill('solid', fgColor='F8CBAD'),
}

THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='top')


# ---------------------------------------------------------------------------
# Document data — file-level granularity matching Hugin A delivery, no dates
# (id, document, description / scope, primary owner, RFQ / source)
# ---------------------------------------------------------------------------

SECTIONS = [
    # ============================================================
    # PROJECT-LEVEL — Sifab planning, control and reporting
    # ============================================================
    ('A. Project planning, control & reporting (Sifab-led)', [
        ('A.1', 'Master schedule / Production plan',
         'Schedule from PO acceptance through FAT, packing, shipment, '
         'disassembly, offshore re-assembly, SAT and close-out.',
         'HW + SF', 'RFQ §1.16-1'),
        ('A.2', 'Quality & Inspection Plan (QIP)',
         'Inspection points (R/H/W), witness/hold notification scheme, '
         'acceptance criteria.',
         'HW + SF', 'RFQ §1.16-2'),
        ('A.3', 'Vendor / Supplier Document List (SDL)',
         'Master document register for the project. Sent to MDCC. '
         '(Hugin A reference: SIF-525-001).',
         'SF', 'RFQ §1.16-3 / Hugin A 01'),
        ('A.4', 'Monthly progress report',
         'Issued monthly to Anne + Martin using Guidant template. '
         '(Hugin A reference: SIF-525-003).',
         'SF', 'PO note / Hugin A 03'),
        ('A.5', 'Vendor Sub Supplied Items – Status xlsx',
         'Per Guidant template — tracks all sub-suppliers (HW, Pemberton, '
         'Flowtec, paint shop, Blu Electro, Intertek, Håkull). Issued with '
         'each progress report.',
         'SF', 'PO template'),
        ('A.6', 'Vendor Milestone Completion Certificate',
         'Issued by Sifab at each Milestone (M1–M4) using Guidant template, '
         'subject to Guidant approval before invoicing.',
         'SF', 'PO MS section'),
        ('A.7', 'Cost Progress Certificate (per milestone)',
         'Cost progress aligned with milestones for invoicing. '
         '(Hugin A reference: TFMC Cost Progress Certificate).',
         'SF', 'standard / Hugin A'),
        ('A.8', 'Project close-out report',
         'Project completion summary, document index, lessons learned.',
         'SF', 'standard'),
        ('A.9', 'Document transmittal (Transmittel)',
         'Cover sheet for each batch of documents sent to Guidant / Honeywell.',
         'SF', 'standard'),
    ]),

    # ============================================================
    # 1. DRAWINGS & DOCUMENTS  (Hugin A package 1)
    # ============================================================
    ('1. Drawings & Documents', [
        ('1.1.a', 'General Arrangement (GA) drawing — full SVP skid',
         'Plan + elevation views, key dimensions, weights, COG, lifting points, '
         'interface list (process, electrical, instrument).',
         'HW', 'RFQ §1.16-4 / Hugin A 1.1'),
        ('1.1.b', 'Name plate drawing',
         'SS316 nameplate with CE, ATEX, PED, tag number details. '
         '(Hugin A reference: 1.1 Name Plate).',
         'HW', 'RFQ §1.14 / Hugin A 1.1'),
        ('1.2', 'Service Clearance Prover drawing',
         'Maintenance access envelope around prover.',
         'HW', 'Hugin A 1.2'),
        ('1.3', 'Frame Mounting (base frame) GA drawing',
         'Base frame mounting interface to platform foundation.',
         'HW', 'Hugin A 1.3'),
        ('1.4.a', 'Base Structure Frame — Index',
         'Index of base structure / frame structural drawings + reports.',
         'HW', 'Hugin A 1.4'),
        ('1.4.b', 'Base Structure Frame — Reports',
         'Structural calculations + reports for the frame.',
         'HW', 'Hugin A 1.4'),
        ('1.5.a', 'Paint Procedure',
         'Coating procedure used for HW factory paint (motor, gearbox).',
         'HW', 'DEV-004 / Hugin A 1.5'),
        ('1.5.b', 'Paint Certification',
         'Coating certificate for HW factory paint.',
         'HW', 'DEV-004 / Hugin A 1.5'),
        # Snorre A-specific drawings
        ('1.6', 'Split-module drawings (modular split)',
         'Per-module dimensions, dry/operating weights, COG, lifting lugs '
         'per NORSOK R-002, handling description. Snorre A-specific. '
         'Drives DEV-007.',
         'HW + SF', 'RFQ §1.7 / DEV-007'),
        ('1.7', '3D model (.stp / native CAD)',
         'For Guidant integration with Snorre A platform model.',
         'HW', 'RFQ §1.16-4'),
        ('1.8', 'P&ID',
         'Process and instrument diagram for the SVP system.',
         'HW', 'standard'),
        ('1.9', 'Wiring diagram',
         'Termination diagram, JB layouts, cable schedule, grounding scheme. '
         '(Hugin A reference: SIF-525-009).',
         'HW', 'RFQ §1.16-10 / Hugin A 09 / lesson C-PM-04'),
        ('1.10', 'GA drawing — Water draw panel',
         'Sifab/Flowtec panel: solenoid, manual valves, instrument hookups.',
         'SF', 'RFQ §1.16-5'),
        ('1.11', 'GA drawing — Seraphin can + cabinet',
         'Pemberton can in SS316 cabinet, lifting/transport details.',
         'SF', 'RFQ §1.16-6 / DEV-017'),
        ('1.12', 'GA drawing — Wooden crate(s) for shipment',
         'Crate design with forklift / crane lift points (per quote Pos 1).',
         'HW', 'quote Pos 1'),
    ]),

    # ============================================================
    # 2. SVP TEST PROCEDURES & CERTIFICATES (Hugin A package 2)
    # ============================================================
    ('2. SVP Test Procedures & Certificates', [
        ('2.1', 'SVP Traceability Documentation',
         'Calibration traceability chain to NIST. (Hugin A 2.1).',
         'HW', 'RFQ §1.12 / Hugin A 2.1'),
        ('2.2.a', 'FAT / Calibration procedure',
         'Gravimetric calibration, water draw test with Seraphin, repeatability test.',
         'HW + SF', 'RFQ §1.16-14'),
        ('2.2.b', 'FAT / Calibration test report',
         'Witnessed calibration result, repeatability data, traceability statement. '
         '(Hugin A 2.2).',
         'HW', 'RFQ §1.16-27 / Hugin A 2.2'),
        ('2.3', 'Instrument Calibration Certificates (FAT instruments)',
         'Traceable calibration certs for instruments used during FAT. '
         '(Hugin A 2.2).',
         'HW', 'RFQ §1.12 / Hugin A 2.2'),
        ('2.4', 'Pressure test procedure',
         'Hydrostatic + leak test sequence per PED / API MPMS 4.2.',
         'HW', 'RFQ §1.16-16'),
        ('2.5', 'Pressure test report',
         'Hydrostatic + leak test record, including 1.5× BD20X (DEV-016).',
         'HW', 'RFQ §1.16 / DEV-016'),
        ('2.6', 'Hydrotest pressure calculation',
         'Per BD20X (1.5 × design pressure ≈ 73.5 barg).',
         'HW', 'DEV-016'),
        # Snorre A-specific testing
        ('2.7', 'SAT / Water draw procedure (Snorre A)',
         'Re-assembly verification, water draw on platform, repeatability '
         '≤ 0.020% demonstration. Drives DEV-008.',
         'HW + SF', 'RFQ §1.16-15 / DEV-008'),
        ('2.8', 'SAT / Water draw test report (Snorre A)',
         'Post re-assembly verification, water draw vs Seraphin, '
         '≤ 0.020% repeatability demonstration.',
         'HW + SF', 'RFQ §1.16-28 / DEV-008'),
        ('2.9', 'Disassembly + packing procedure',
         'Procedure for disassembly at Norwegian fab yard, packing in wooden '
         'crates, transport. Honeywell supervisor lead.',
         'HW + SF', 'quote Pos 4'),
        ('2.10', 'Offshore re-assembly procedure',
         'Lifting, fitting, alignment, flange bolting, cable re-connect, '
         'pressure / leak test sequence on Snorre A.',
         'HW + SF', 'quote Pos 4'),
    ]),

    # ============================================================
    # 3.a FLOW TUBE — WELDING (Hugin A package 3.a)
    # ============================================================
    ('3.a Flow Tube — Welding & NDE', [
        ('3.1.a', 'Flow Tube Certification Package — Index',
         'Index document for the flow tube certification pack. (Hugin A 3.1).',
         'HW', 'RFQ §1.16'),
        ('3.1.b', 'Vendor — Flow Tube Certification Checklist',
         'Pre-shipment checklist with sign-offs.',
         'HW', 'Hugin A 3.1'),
        ('3.2', 'Casting / Material Test Certificate (flow tube)',
         'EN 10204 3.1 material certificate for flow tube body.',
         'HW', 'RFQ §1.16-24 / Hugin A 3.2'),
        ('3.3.a', 'MTC — Anvilet / Weld-O-Let',
         'Material certs for anvilet/weld-o-let connections.',
         'HW', 'Hugin A 3.3'),
        ('3.3.b', 'MTC — Elbows',
         'Material certs for elbow components.',
         'HW', 'Hugin A 3.3'),
        ('3.3.c', 'MTC — Extension Pipe',
         'Material certs for extension pipes.',
         'HW', 'Hugin A 3.3'),
        ('3.3.d', 'MTC — In/Out/Bleed/Drain Flanges',
         'Material certs for inlet, outlet, bleed and drain flanges.',
         'HW', 'Hugin A 3.3'),
        ('3.4.a', 'Certified Material Test Report — Weld Rod',
         'Weld consumable certs.',
         'HW', 'Hugin A 3.4'),
        ('3.4.b', 'Liquid Penetrant Test (PT) report',
         'Surface examination report.',
         'HW', 'Hugin A 3.4'),
        ('3.4.c', 'Radiographic Test (RT) report',
         'Volumetric weld examination report.',
         'HW', 'Hugin A 3.4'),
        ('3.4.d', 'NDE / NDT certifications — Technician certs',
         'Level II/III qualifications for NDE operators. '
         '(Hugin A 3.4).',
         'HW', 'RFQ §1.16-11 / Hugin A 3.4'),
        ('3.4.e', 'Weld Map',
         'Weld identification map covering all pressure welds.',
         'HW', 'RFQ §1.16-19 / Hugin A 3.4'),
        ('3.4.f', 'Welder Certifications (incl. WPS, PQR, DNV procedures)',
         'Welder qualifications, WPS and PQR per ASME IX (DEV-003). '
         '(Hugin A 3.4).',
         'HW', 'RFQ §1.16-18 / DEV-003 / Hugin A 3.4'),
        ('3.5.a', 'Certificate of Compliance — Plating (chrome)',
         'Hard chrome plating certificate for flow tube bore.',
         'HW', 'Hugin A 3.5'),
        ('3.5.b', 'Certificate of Conformance (flow tube)',
         'Manufacturer conformance statement.',
         'HW', 'Hugin A 3.5'),
        ('3.5.c', 'Dimensional Inspection Report — TruStop',
         'Dimensional inspection of flow tube at TruStop facility.',
         'HW', 'Hugin A 3.5'),
        ('3.5.d', 'Dimensional Inspection Report — Vendor',
         'Dimensional inspection at flow tube vendor (India).',
         'HW', 'Hugin A 3.5'),
        ('3.5.e', 'TruStop PMI Report',
         'Positive Material Identification at TruStop. (Hugin A 3.5).',
         'HW', 'RFQ §1.16-25 / DEV-010 / Hugin A 3.5'),
        ('3.5.f', 'Vendor PMI Report',
         'PMI at flow tube vendor.',
         'HW', 'RFQ §1.16-25 / DEV-010 / Hugin A 3.5'),
        ('3.6', 'Welding Procedure Specification (WPS)',
         'For all pressure-containing welds. ASME IX per DEV-003.',
         'HW', 'RFQ §1.16-18 / DEV-003'),
        ('3.7', 'Welding Procedure Qualification Record (WPQ / PQR)',
         'Backup data for the WPS.',
         'HW', 'RFQ §1.16-18 / DEV-003'),
        ('3.8', 'NDE procedure',
         'Non-destructive examination of welds (RT/UT/PT/MT as applicable).',
         'HW', 'RFQ §1.16-13'),
        ('3.9', 'Weld record sheet (per joint)',
         'WPS reference, welder ID, NDE results.',
         'HW', 'RFQ §1.16-20'),
    ]),

    # ============================================================
    # 3.b MAJOR WETTED PARTS & COMPONENTS (Hugin A package 3.b)
    # ============================================================
    ('3.b Major Wetted Parts & Components', [
        ('3.6.a', 'Major Wetted Parts & Components — Traceability and Cert. Index',
         'Index document.',
         'HW', 'Hugin A 3.6'),
        ('3.6.b', 'Major Wetted Parts Traceability and Certification (full pack)',
         'EN 10204 3.1 certs for all wetted parts.',
         'HW', 'RFQ §1.16-24 / Hugin A 3.6'),
        ('3.7', 'Valves traceability + certs (incl. DBB, vent)',
         'Material + functional certs for all process valves.',
         'HW', 'Hugin A 3.7'),
        ('3.8', 'NACE MR0175 / ISO 15156-3 compliance certificates',
         'For all wetted parts in sour service. Critical small-bore items.',
         'HW + SF', 'RFQ §1.8-3 / lesson C-HSE-02'),
        ('3.9', '6Mo material certs (Sifab free-issue tubing/fittings)',
         'For tubing and fittings per ST701/SF712.',
         'SF', 'DEV-009 / lesson Aker BP'),
    ]),

    # ============================================================
    # 4. SVP PROVER SYSTEM CERTIFICATIONS & CONFORMITIES (pkg 4)
    # ============================================================
    ('4. SVP Prover System Certifications & Conformities', [
        ('4.1.a', 'Certificate of Conformance for Prover',
         'Honeywell Cert. of Conformance for full prover. (Hugin A 4.1).',
         'HW', 'RFQ §1.16'),
        ('4.1.b', 'Honeywell Declaration of Conformity',
         'Full CE / EU Declaration of Conformity covering PED + ATEX + Machinery + EMC.',
         'HW', 'RFQ §1.0-11 / §1.16-23 / Hugin A 4.1'),
        ('4.2', 'Safety Instructions',
         'Safety instructions for installation and operation.',
         'HW', 'Hugin A 4.2'),
        ('4.3', 'PED 2014/68/EU certificate',
         'Pressure Equipment Directive certification.',
         'HW', 'RFQ §1.1-3'),
        ('4.4', 'ATEX certificates (mechanical + electrical, complete pack)',
         'Per component (motor, controller, switches, JBs). '
         '(Hugin A reference: SIF-525-005 + SVP-ATEX-HON160001).',
         'HW', 'RFQ §1.16-17 / Hugin A 05'),
        ('4.5', 'Machinery Directive 2006/42/EC compliance documentation',
         'Machinery Directive coverage (CR per RFQ §1.1-5).',
         'HW', 'RFQ §1.1-5'),
        ('4.6', 'EMC Directive 2014/30/EU compliance documentation',
         'EMC Directive coverage (CR per RFQ §1.1-6).',
         'HW', 'RFQ §1.1-6'),
        ('4.7', 'API MPMS 4.2 compliance statement',
         'Compliance with displacement prover standard.',
         'HW', 'RFQ §1.1-1'),
        ('4.8', 'Justervesenet certificate (Seraphin can)',
         '3rd-party traceable volume certification.',
         'SF/Sub', 'RFQ §1.0-3 / DEV-017'),
    ]),

    # ============================================================
    # 5. SVP CONTROLLER & TRANSMITTER (Hugin A package 5)
    # ============================================================
    ('5. SVP Controller & Transmitter Documentation', [
        ('5.1', 'SVP Controller — Certificate of Conformance',
         'CoC for the SVP controller. (Hugin A 5.1).',
         'HW', 'Hugin A 5.1'),
        ('5.2.a', 'Control Drawings',
         'Honeywell control drawings for the controller.',
         'HW', 'Hugin A 5.2'),
        ('5.2.b', 'Controller CD / digital deliverables',
         'Configuration files, manuals, software (Hugin A had a CD).',
         'HW', 'Hugin A 5.2'),
        ('5.3', 'Pressure Transmitter Amendment Letter',
         'For Sifab free-issue transmitters (per RFQ).',
         'SF', 'Hugin A 5.3 / RFQ §1.6-8'),
        ('5.4', 'Transmitter Calibration Reports',
         'Calibration certs for pressure transmitters.',
         'SF', 'Hugin A 5.4'),
        ('5.5', 'Product datasheet — controller + transmitters',
         '(Hugin A reference: SIF-525-009).',
         'HW + SF', 'Hugin A 09'),
        ('5.6', 'Cable schedule (as-built)',
         'All cables: from/to, length, type, gland, tag.',
         'HW + SF', 'standard'),
        ('5.7', 'Termination schedule (as-built)',
         'JB and instrument terminal connections.',
         'HW + SF', 'standard'),
        ('5.8', 'Instrument index (as-built)',
         'List of all instruments with tag, range, calibration, location.',
         'HW + SF', 'standard'),
    ]),

    # ============================================================
    # 6. MOTOR DOCUMENTS (Hugin A package 6)
    # ============================================================
    ('6. Motor Documents', [
        ('6.1.a', 'Speed Reducer (gearbox) — Catalogue',
         'Catalogue / spec for speed reducer.',
         'HW', 'Hugin A 6.1'),
        ('6.1.b', 'Speed Reducer (gearbox) — Manual',
         'Operation + maintenance manual for gearbox.',
         'HW', 'Hugin A 6.1'),
        ('6.2', 'Motor Amendment Letter',
         'Letter formalising motor model / spec changes (e.g. 60 Hz per DEV-002).',
         'HW', 'Hugin A 6.2 / DEV-002'),
        ('6.3', 'Motor specification + datasheet',
         '230 VAC, 3-ph, 60 Hz, ATEX Ex de or Ex e (DEV-002).',
         'HW', 'RFQ §1.5 / DEV-002'),
        ('6.4', 'Motor ATEX certificate',
         'Standalone ATEX cert for motor.',
         'HW', 'RFQ §1.16-17'),
        ('6.5', 'Motor test report / FAT',
         'Motor factory test results.',
         'HW', 'standard'),
        ('6.6', 'Motor painting / coating documentation',
         'HW factory offshore paint or NORSOK exemption (DEV-004).',
         'HW', 'DEV-004'),
    ]),

    # ============================================================
    # 7. OTHER PROJECT-SPECIFIC EQUIPMENT (Hugin A package 7)
    # ============================================================
    ('7. Other Project-Specific Equipment Documents', [
        ('7.1', 'Temperature RTD Sensor Calibration Reports',
         'For Sifab free-issue RTD sensors. (Hugin A 7.1).',
         'SF', 'Hugin A 7.1 / RFQ §1.6'),
        ('7.2', 'Cable Glands Amendment Letter',
         'For Sifab free-issue cable glands. (Hugin A 7.2).',
         'SF', 'Hugin A 7.2 / DEV-011'),
        ('7.3', 'Enclosure Box Amendment Letter',
         'For SS316 enclosure boxes. (Hugin A 7.3).',
         'SF', 'Hugin A 7.3 / RFQ §1.8-6'),
        ('7.4', 'Optical switch certificates',
         'KIT OPTICAL SWITCH set of 3 (P/N 44110148).',
         'HW', 'quote Pos 5'),
        ('7.5', 'Motor stop switch certificate',
         'Motor stop switch — O models (P/N 44107515).',
         'HW', 'quote Pos 5'),
        ('7.6', 'Seal Kit certificate',
         'SEAL KIT CSK S85 150/300/600# CRB (P/N 44105627).',
         'HW', 'quote Pos 5'),
        ('7.7', 'Thermowells — material certs + dim. spec',
         '3× process + 1× rod thermowells, ½″ NPTF, BD20X.',
         'SF', 'RFQ §1.6-5,6'),
        ('7.8', 'B&B (DBB) valve certs',
         '½″ DB&B per VDS-MHBD102R / TR2000.',
         'SF', 'RFQ §1.6-7'),
        ('7.9', 'Wake frequency calculations (process thermowells)',
         'Per ASME PTC 19.3 TW.',
         'SF', 'RFQ §1.16-9'),
        ('7.10', 'Lifting calculations + certs (lugs, slings, shackles)',
         'Per NORSOK R-002. Drives DEV-014.',
         'SF', 'RFQ §1.13 / DEV-014'),
        ('7.11', 'Lifting procedure (offshore)',
         'Snorre A platform crane interface.',
         'SF', 'RFQ §1.13'),
    ]),

    # ============================================================
    # 8. SVP OPERATION & SERVICE MANUAL (Hugin A package 8)
    # ============================================================
    ('8. SVP Operation & Service Manual', [
        ('8.1', 'Honeywell Enraf SVP Operation & Service Manual (IOM)',
         'Full IOM manual for operation and maintenance. (Hugin A 8.0).',
         'HW', 'RFQ §1.16-21,22 / Hugin A 8.0'),
        ('8.2', 'Operating manual extract / quick guide',
         'Operating quick reference, normal/abnormal conditions, calibration interval.',
         'HW', 'RFQ §1.16-21'),
        ('8.3', 'Maintenance manual extract',
         'Routine + corrective maintenance, parts identification, lubrication.',
         'HW', 'RFQ §1.16-22'),
        ('8.4', 'Storage and preservation instructions',
         'Post-FAT storage at Sifab + during transit + on Snorre A pre-installation.',
         'HW + SF', 'standard'),
    ]),

    # ============================================================
    # 9. NORSOK / TR DATASHEETS (Snorre A specific)
    # ============================================================
    ('9. NORSOK datasheets (Snorre A specific)', [
        ('9.1', 'NORSOK datasheet — prover (M-CR-630 / MDS)',
         'Material Data Sheet per NORSOK M-630 for all wetted/pressure parts.',
         'HW + SF', 'RFQ §1.16-8 / DEV-009'),
        ('9.2', 'NORSOK datasheet — thermowells',
         'Material + dimensional spec per TR2000 BD20X.',
         'SF', 'RFQ §1.16-8 / RFQ §1.6'),
        ('9.3', 'NORSOK datasheet — motor',
         'Motor spec, ATEX, IP rating, power requirements.',
         'HW', 'RFQ §1.16-8'),
        ('9.4', 'NORSOK datasheet — switches (optical, motor stop)',
         'Spec for optical switches and motor stop switch.',
         'HW', 'RFQ §1.16-8'),
        ('9.5', 'NORSOK datasheet — water draw kit',
         'Spec for water draw kit (Flowtec).',
         'SF', 'RFQ §1.16-8'),
    ]),

    # ============================================================
    # 10. NORSOK PAINTING — Sifab in-Norway scope
    # ============================================================
    ('10. NORSOK painting (Sifab in-Norway scope)', [
        ('10.1', 'Surface treatment procedure (NORSOK M-501 System 6C)',
         'Applied by Norwegian paint shop after FAT. Polyurethane-free per Equinor. '
         'Submitted within 2 WAO for pre-approval.',
         'SF', 'RFQ §1.16-7 / DEV-004 / lesson C-COAT-03'),
        ('10.2', 'Surface treatment + coating report',
         'Includes DFT, adhesion, holiday testing per M-501.',
         'SF', 'RFQ §1.16-30'),
        ('10.3', 'Painting / coating qualification (operators + system)',
         'NORSOK M-501 §10.2.3 operator qualification.',
         'SF/Sub', 'lesson C-COAT-01'),
        ('10.4', 'Coating SDS / paint product approval (Equinor TR0042)',
         'Paint product compliance with TR0042. Polyurethane-free.',
         'SF/Sub', 'lesson C-COAT-04'),
    ]),

    # ============================================================
    # 11. SPARE PARTS & SERVICES
    # ============================================================
    ('11. Spare parts & services', [
        ('11.1', 'SPIR (Spare Parts and Interchangeability Record)',
         'Sifab/Honeywell with prices, lead times, part numbers, GA cross-references. '
         '(Hugin A reference: SIF-525-017).',
         'HW + SF', 'RFQ §1.16-20 / Hugin A 17'),
        ('11.2', 'Recommended 2-year operating spares list',
         'Per Honeywell standard (optical switches, motor stop switch, seal kit).',
         'HW', 'RFQ §1.15 / quote Pos 5'),
        ('11.3', 'Commissioning spares list',
         'Single-use spares for commissioning phase.',
         'HW', 'RFQ §1.15'),
    ]),

    # ============================================================
    # 12. PRESERVATION (Hugin A had separate doc — SIF-525-028)
    # ============================================================
    ('12. Preservation', [
        ('12.1', 'Preservation procedure',
         'For storage at Sifab between FAT and offshore re-assembly + during transit. '
         '(Hugin A reference: SIF-525-028).',
         'SF', 'RFQ §1.16 / Hugin A 28'),
        ('12.2', 'Preservation report',
         'Records preservation actions during storage / transit.',
         'SF', 'RFQ §1.16-32'),
    ]),

    # ============================================================
    # 13. PUNCH LISTS
    # ============================================================
    ('13. Punch lists', [
        ('13.1', 'Pre-FAT punch list (Honeywell self-assessment)',
         'Honeywell to complete BEFORE inviting Sifab to FAT (per lesson C-QM-03).',
         'HW', 'lesson C-QM-03'),
        ('13.2', 'FAT punch list (Sifab + Guidant)',
         'Outstanding items at FAT, with close-out plan.',
         'HW + SF', 'RFQ §1.12'),
        ('13.3', 'SAT punch list (Snorre A)',
         'Outstanding items after offshore re-assembly + SAT.',
         'HW + SF', 'RFQ §1.12'),
    ]),

    # ============================================================
    # 14. LOGISTICS / TRANSPORT / WEIGHING
    # ============================================================
    ('14. Logistics, transport & weighing', [
        ('14.1', 'Weighing report (per module, dry + operating)',
         'Critical for modular split + Snorre A crane capacity.',
         'HW + SF', 'RFQ §1.16-31 / DEV-007'),
        ('14.2', 'Packing list (per crate)',
         'Per Honeywell packaging code 01 — wooden crate.',
         'HW', 'quote Pos 1'),
        ('14.3', 'Cable protection inspection (pre-shipment)',
         'Per lesson C-EI-07 — secure cables, SS316 clamps at 300 mm.',
         'HW', 'lesson C-EI-07'),
        ('14.4', 'Shipping documentation (FCA Sandnes)',
         'Bill of lading, customs, export docs.',
         'HW + SF', 'PO Incoterms'),
        ('14.5', 'Lifting plan + offshore transport plan',
         'Sea/road transport Sifab → Snorre A quayside (Håkull AS).',
         'SF/Sub', 'standard'),
    ]),

    # ============================================================
    # 15. CONFIGURATION & DATA SHEETS (with delivery)
    # ============================================================
    ('15. Configuration & as-built data sheets', [
        ('15.1', 'Final equipment datasheet — as-built (QCF524 or equiv.)',
         'Honeywell as-built configuration sheet.',
         'HW', 'standard'),
        ('15.2', 'Configuration sheets (Hugin A reference: package 7)',
         'Customer-specific configuration items.',
         'HW + SF', 'Hugin A 7'),
        ('15.3', 'Tag schedule (final, all instruments)',
         'Confirmed tag numbers from Guidant, applied to nameplates.',
         'HW + SF', 'lesson C-PM-03'),
    ]),
]


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = 'Document List'

# Title
ws['A1'] = 'SP-01415 Snorre A SVP085 — Document List for Honeywell Review'
ws['A1'].font = TITLE_FONT
ws['A1'].fill = TITLE_FILL
ws.merge_cells('A1:I1')
ws.row_dimensions[1].height = 28

# Project header
header_row_start = 2
project_meta = [
    ('Project', 'SP-01415 Snorre A Small Volume Prover (SVP085)'),
    ('Customer Project', 'GM-5341 Snorre A Oil Metering'),
    ('Customer', 'Guidant — Measurement Solutions Norway AS'),
    ('End Client', 'Equinor — Snorre A platform'),
    ('Honeywell Proposal', '10465986-O-1010834 R0'),
    ('Sifab PO (received)', '4500998501 dated 29 April 2026'),
    ('Reference projects', 'SP-00525 Hugin A (SVP050) + SP-00577 Valhall (SVP015) — full delivery package mirrored'),
    ('Document', 'Document List for review with Honeywell — no dates yet'),
    ('Prepared by', 'Sifab AS'),
    ('Revision', 'Working draft — 5 May 2026'),
]
for i, (k, v) in enumerate(project_meta):
    r = header_row_start + i
    a = ws.cell(row=r, column=1, value=k)
    b = ws.cell(row=r, column=2, value=v)
    a.font = Font(bold=True)
    a.alignment = WRAP
    b.alignment = WRAP

note_row = header_row_start + len(project_meta) + 1
ws.cell(row=note_row, column=1, value='How to use').font = SECTION_FONT
ws.cell(row=note_row, column=1).fill = SECTION_FILL
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=9)

instructions = [
    '1. Honeywell to mark "In HW Std Package?" with Y / N / Partial for each line.',
    '2. Honeywell to add comments where scope is unclear or document is non-standard (extra cost).',
    '3. Sifab will add Sifab comments + propose final delivery schedule (WAO / WD) after this round.',
    '4. Document numbering will be re-assigned per agreed final list before transmittal to Guidant (MDCC).',
    '5. Sections 1–8 mirror the actual Hugin A delivery package (SP-00525) at file-level granularity.',
    '6. Sections 9–15 cover Snorre A-specific additions (modular split, NORSOK painting, SAT, free-issue items).',
    '7. Owner key: HW = Honeywell scope; SF = Sifab scope; HW + SF = combined; SF/Sub or HW/Sub = via sub-supplier.',
]
for i, txt in enumerate(instructions):
    ws.cell(row=note_row + 1 + i, column=1, value=txt).alignment = WRAP
    ws.merge_cells(start_row=note_row + 1 + i, start_column=1, end_row=note_row + 1 + i, end_column=9)

# Column header row
table_header_row = note_row + len(instructions) + 2
columns = [
    '#',
    'Document',
    'Description / scope',
    'Primary Owner',
    'RFQ / Source',
    'In HW Std Package? (Y/N/Partial)',
    'Honeywell Comments',
    'Sifab Comments',
    'Status',
]
for c_idx, name in enumerate(columns, start=1):
    cell = ws.cell(row=table_header_row, column=c_idx, value=name)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = HEADER_ALIGN
    cell.border = BORDER
ws.row_dimensions[table_header_row].height = 32

# Data rows
current_row = table_header_row + 1
for section_title, items in SECTIONS:
    cell = ws.cell(row=current_row, column=1, value=section_title)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(vertical='center')
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
    current_row += 1

    for doc_id, doc_name, desc, owner, rfq_ref in items:
        ws.cell(row=current_row, column=1, value=doc_id)
        ws.cell(row=current_row, column=2, value=doc_name)
        ws.cell(row=current_row, column=3, value=desc)
        ws.cell(row=current_row, column=4, value=owner)
        ws.cell(row=current_row, column=5, value=rfq_ref)
        ws.cell(row=current_row, column=6, value='')
        ws.cell(row=current_row, column=7, value='')
        ws.cell(row=current_row, column=8, value='')
        ws.cell(row=current_row, column=9, value='Open')

        for col_idx in range(1, 10):
            c = ws.cell(row=current_row, column=col_idx)
            c.alignment = WRAP
            c.border = BORDER

        owner_cell = ws.cell(row=current_row, column=4)
        if owner in OWNER_FILLS:
            owner_cell.fill = OWNER_FILLS[owner]
            owner_cell.font = Font(bold=True)

        current_row += 1

# Column widths
widths = [10, 50, 65, 14, 28, 16, 32, 32, 12]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = f'B{table_header_row + 1}'

wb.save(OUTPUT)
total = sum(len(items) for _, items in SECTIONS)
print(f'Saved: {OUTPUT}')
print(f'Sections: {len(SECTIONS)} | Documents: {total}')
