"""Build SP-01415 Snorre A Document Status & Tracking workbook.

Mirrors Oliver Vetland's Hugin A tracker structure:
"Dokument status og oppfølging.xlsx" — one row per transmittal-level document
with submission tracking. Document numbering: SIF-1415-001 → SIF-1415-NNN.

The companion script tools/sp01415_scan_documents.py reads this workbook,
scans email + project folders, and updates dates/status in-place.

Saved at the project root for visibility:
    Zigma360 / Projects / SP-01415 .../ SP-01415_Dokument_status_og_oppfølging.xlsx
"""

import os
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHARED = Path(os.environ['USERPROFILE']) / 'OneDrive - Sifab AS' / 'Dokumenter - Felles'
PROJECT = SHARED / 'Zigma360' / 'Projects' / 'SP-01415 Small Volume Prover Snorre A'
OUTPUT = PROJECT / 'SP-01415_Dokument_status_og_oppfølging.xlsx'

# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------
TITLE_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE_FONT = Font(bold=True, color='FFFFFF', size=14)

HEADER_FILL = PatternFill('solid', fgColor='2E75B6')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

OWNER_FILLS = {
    'Honeywell': PatternFill('solid', fgColor='F8CBAD'),
    'Sifab AS': PatternFill('solid', fgColor='C6E0B4'),
    'Sifab AS / HW': PatternFill('solid', fgColor='FFE699'),
    'Guidant': PatternFill('solid', fgColor='BDD7EE'),
    'Pemberton': PatternFill('solid', fgColor='C6E0B4'),
    'Flowtec': PatternFill('solid', fgColor='C6E0B4'),
    'Paint shop': PatternFill('solid', fgColor='C6E0B4'),
}

STATUS_FILLS = {
    'Open': PatternFill('solid', fgColor='F4B084'),
    'In progress': PatternFill('solid', fgColor='FFD966'),
    'Submitted': PatternFill('solid', fgColor='9DC3E6'),
    'In review': PatternFill('solid', fgColor='FFD966'),
    'Approved': PatternFill('solid', fgColor='A9D08E'),
    'Code 1': PatternFill('solid', fgColor='A9D08E'),
    'Code 2': PatternFill('solid', fgColor='FFD966'),
    'Code 3': PatternFill('solid', fgColor='F4B084'),
    'Code 4': PatternFill('solid', fgColor='FF7C80'),
    'Closed': PatternFill('solid', fgColor='A9D08E'),
}

THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='top')


# ---------------------------------------------------------------------------
# Document register — Oliver-style transmittal level
# (id, title, submit_retain, due_wao_or_wd, owner, notes)
# WAO = Weeks After Order, WD = With Delivery, AD = After Drawing approved
# ---------------------------------------------------------------------------
PO_DATE = date(2026, 4, 30)

REGISTER = [
    # Project planning & control
    ('SIF-1415-001', 'Supplier Document List (SDL)', 'S', '2 WAO', 'Sifab AS',
     'Master document register transmitted to MDCC. (Hugin A: SIF-525-001).'),
    ('SIF-1415-002', 'Inspection & Test Plan (ITP / QIP)', 'S', '4 WAO', 'Sifab AS / HW',
     'Combined Sifab + Honeywell QIP. (Hugin A: SIF-525-002).'),
    ('SIF-1415-003', 'Progress Report (Monthly)', 'S', 'Monthly from kick-off', 'Sifab AS',
     'First report due 2026-05-08 per PO. Guidant template. (Hugin A: SIF-525-003).'),
    ('SIF-1415-004', 'NORSOK Instrument Datasheet — Prover, Motor, Switches', 'S', '4 WAO', 'Sifab AS / HW',
     '(Hugin A: SIF-525-004).'),
    # Drawings & calculations
    ('SIF-1415-005', 'Sizing and design calculations', 'S', 'AD', 'Honeywell',
     '(Hugin A: SIF-525-006).'),
    ('SIF-1415-006', 'Wake Frequency Calculations for Thermowells', 'S', '4 WAO', 'Sifab AS',
     '(Hugin A: SIF-525-007).'),
    ('SIF-1415-007', 'GA drawing — Prover (full SVP skid)', 'S', 'AD', 'Honeywell',
     '(Hugin A: SIF-525-008).'),
    ('SIF-1415-008', 'GA drawing — Thermowells', 'S', '4 WAO', 'Sifab AS',
     '(Hugin A: SIF-525-029).'),
    ('SIF-1415-009', 'GA drawing — Seraphin can + cabinet', 'S', '4 WAO', 'Sifab AS',
     '(Hugin A: SIF-525-031). SS316 cabinet.'),
    ('SIF-1415-010', 'GA drawing — Water draw panel', 'S', '4 WAO', 'Flowtec',
     'Sifab/Flowtec scope.'),
    ('SIF-1415-011', 'Wiring diagram / Product datasheet (Motor and Switches)', 'S', '4 WAO', 'Honeywell',
     '(Hugin A: SIF-525-009).'),
    # Snorre A specific drawings
    ('SIF-1415-012', 'Split-module drawings + lifting plan (modular split)', 'S', '4 WAO', 'Sifab AS / HW',
     'Per-module dim + weight + COG, lifting per NORSOK R-002. Drives DEV-007.'),
    ('SIF-1415-013', '3D model (.stp) — full SVP skid', 'S', '4 WAO', 'Honeywell',
     'For Guidant Snorre A platform integration.'),
    # Procedures
    ('SIF-1415-014', 'FAT / Calibration procedure', 'S', 'AD', 'Honeywell',
     '(Hugin A: SIF-525-010).'),
    ('SIF-1415-015', 'Hydrostatic pressure test procedure', 'S', '2 WAO', 'Sifab AS / HW',
     '(Hugin A: SIF-525-011).'),
    ('SIF-1415-016', 'PMI procedure', 'S', '4 WAO', 'Honeywell',
     'Clarify "less hardware" scope. (Hugin A: SIF-525-012). DEV-010.'),
    ('SIF-1415-017', 'NDT procedure', 'S', 'AD', 'Honeywell',
     '(Hugin A: SIF-525-013).'),
    ('SIF-1415-018', 'NDT Operator certificates', 'S', 'AD', 'Honeywell',
     '(Hugin A: SIF-525-014).'),
    ('SIF-1415-019', 'Surface Treatment Procedure (NORSOK M-501 6C)', 'S', 'AD', 'Sifab AS',
     'Sifab in-Norway scope. Polyurethane-free. (Hugin A: SIF-525-015). DEV-004.'),
    ('SIF-1415-020', 'Welding documentation — WPS, WPQR, Welder certificates', 'S', 'AD', 'Honeywell',
     'ASME IX per DEV-003. (Hugin A: SIF-525-016).'),
    # Snorre A specific procedures
    ('SIF-1415-021', 'Disassembly + packing procedure (onshore)', 'S', '8 WAO', 'Sifab AS / HW',
     'Norwegian fab yard, wooden crates, transport. HE supervised.'),
    ('SIF-1415-022', 'Offshore re-assembly procedure (Snorre A)', 'S', '8 WAO', 'Sifab AS / HW',
     'Lifting, fitting, alignment, flange bolt-up, cable re-connect, leak test.'),
    ('SIF-1415-023', 'SAT / Water draw procedure (Snorre A)', 'S', '8 WAO', 'Sifab AS / HW',
     'Repeatability ≤ 0.020% demonstration. DEV-008.'),
    ('SIF-1415-024', 'Lifting plan (offshore) per NORSOK R-002', 'S', '8 WAO', 'Sifab AS',
     'Snorre A platform crane interface. DEV-014.'),
    ('SIF-1415-025', 'Preservation procedure', 'S', '8 WAO', 'Sifab AS',
     '(Hugin A: SIF-525-028).'),
    # SPIR + spares
    ('SIF-1415-026', 'SPIR (Spare Parts and Interchangeability Record)', 'S', '4 WAO', 'Sifab AS / HW',
     '(Hugin A: SIF-525-017).'),
    # Reports & certs (with delivery / after delivery)
    ('SIF-1415-027', 'Material Certificates (EN 10204 3.1, all wetted/pressure parts)', 'S', '4 WAO', 'Honeywell',
     '(Hugin A: SIF-525-018).'),
    ('SIF-1415-028', 'Pressure test report', 'S', 'WD', 'Honeywell',
     '(Hugin A: SIF-525-019). DEV-016.'),
    ('SIF-1415-029', 'PED certificate, EC Declaration of Conformity', 'S', 'WD', 'Honeywell',
     '(Hugin A: SIF-525-020).'),
    ('SIF-1415-030', 'FAT / Calibration Report', 'S', 'WD', 'Honeywell',
     '(Hugin A: SIF-525-021).'),
    ('SIF-1415-031', 'PMI report', 'S', 'WD', 'Honeywell',
     '(Hugin A: SIF-525-022). DEV-010.'),
    ('SIF-1415-032', 'IOM (Installation, Operation and Maintenance) Manual', 'S', 'WD', 'Honeywell',
     '(Hugin A: SIF-525-023).'),
    ('SIF-1415-033', 'NDT reports', 'S', 'WD', 'Honeywell',
     '(Hugin A: SIF-525-024).'),
    ('SIF-1415-034', 'Surface Treatment Report', 'S', 'WD', 'Sifab AS',
     '(Hugin A: SIF-525-025). DFT, adhesion, holiday testing.'),
    ('SIF-1415-035', 'Weld Report', 'S', 'WD', 'Honeywell',
     '(Hugin A: SIF-525-026).'),
    ('SIF-1415-036', 'Meggering reports — cables', 'S', 'WD', 'Honeywell',
     '(Hugin A: SIF-525-027). Lesson C-EI-04: design for individual disconnection.'),
    ('SIF-1415-037', 'Motor documents (spec, ATEX, test, painting)', 'S', '4 WAO', 'Honeywell',
     '(Hugin A: SIF-525-030). DEV-002 motor 60 Hz.'),
    # Snorre A specific reports
    ('SIF-1415-038', 'SAT / Water draw test report (Snorre A)', 'S', 'After SAT', 'Sifab AS / HW',
     'Post offshore re-assembly. DEV-008.'),
    ('SIF-1415-039', 'Weighing report (per module, dry + operating)', 'S', 'WD', 'Sifab AS / HW',
     'Critical for Snorre A crane capacity. DEV-007.'),
    ('SIF-1415-040', 'Preservation report', 'S', 'WD', 'Sifab AS',
     ''),
    ('SIF-1415-041', 'Justervesenet certificate (Seraphin can)', 'S', 'WD', 'Pemberton',
     'DEV-017.'),
    ('SIF-1415-042', 'ATEX certificates pack (JBs, Motor)', 'S', '8 WAO', 'Honeywell',
     '(Hugin A: SIF-525-005). Std incl. SVP-ATEX-HON160001.'),
    # Internal / retain only
    ('SIF-1415-043', 'Pre-FAT punch list (Honeywell self-assessment)', 'R', 'Pre-FAT', 'Honeywell',
     'Honeywell completes BEFORE Sifab attends FAT. Lesson C-QM-03.'),
    ('SIF-1415-044', 'FAT punch list', 'R', 'At FAT', 'Sifab AS / HW',
     'Outstanding items + close-out plan.'),
    ('SIF-1415-045', 'SAT punch list (Snorre A)', 'R', 'At SAT', 'Sifab AS / HW',
     'Outstanding items after offshore re-assembly.'),
    ('SIF-1415-046', 'Vendor Milestone Completion Certificate (M1–M4)', 'S', 'Per milestone', 'Sifab AS',
     'Per Guidant template, required to invoice each milestone.'),
    ('SIF-1415-047', 'Vendor Sub Supplied Items — Status xlsx', 'S', 'Monthly w/ progress', 'Sifab AS',
     'Per Guidant template.'),
    ('SIF-1415-048', 'Deviation List (rev. controlled)', 'S', 'Per revision', 'Sifab AS',
     'Rev 0 = 23 March 2026. Rev 1 = 5 May 2026 (post-PO). Per PO clause.'),
    ('SIF-1415-049', 'Kick-off + pre-production meeting MOM', 'S', 'After meeting', 'Sifab AS',
     'PO requires kick-off and pre-production meeting.'),
    ('SIF-1415-050', 'Project close-out report', 'S', 'After M4', 'Sifab AS',
     'Document index, lessons learned.'),
]


# ---------------------------------------------------------------------------
# Build workbook (only when run as script — never on import)
# ---------------------------------------------------------------------------
def build():
    wb = Workbook()
    ws = wb.active
    ws.title = 'SDL'

    ws['A1'] = 'SP-01415 Snorre A SVP085 — Dokument status og oppfølging'
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = TITLE_FILL
    ws.merge_cells('A1:M1')
    ws.row_dimensions[1].height = 28

    meta = [
        ('Project Name', 'SP-01415 Snorre A SVP085'),
        ('Customer Project', 'GM-5341 Snorre A Oil Metering'),
        ('Customer', 'Guidant — Measurement Solutions Norway AS'),
        ('Supplier', 'Sifab AS'),
        ('Made by', 'Sondre Falch'),
        ('Checked by', ''),
        ('Approved by', ''),
        ('Purchase Order No.', '4500998501'),
        ('PO date', '2026-04-29'),
        ('Date created', '2026-05-05'),
        ('Supplier Rev.', '00'),
    ]
    for i, (k, v) in enumerate(meta):
        r = 2 + i
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
        ws.cell(row=r, column=1).alignment = WRAP
        ws.cell(row=r, column=2).alignment = WRAP

    header_row = 2 + len(meta) + 1
    columns = [
        'Document No.',
        'Document Title',
        'Submit/Retain (S/R)',
        'Planned Submission (WAO/WD)',
        'Date for first Submission',
        'Days overdue first sub',
        'Date for latest Submission',
        'Date for latest Return',
        '# Revisions sent',
        '# Revisions returned',
        'Owner',
        'Status',
        'Filename(s) found in folders',
        'Folder location',
        'Notes',
    ]
    for c_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    ws.row_dimensions[header_row].height = 36

    data_start = header_row + 1
    for i, (doc_id, title, sr, due, owner, notes) in enumerate(REGISTER):
        r = data_start + i
        row_values = [
            doc_id, title, sr, due, '', '', '', '', 0, 0,
            owner, 'Open', '', '', notes,
        ]
        for c_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.alignment = WRAP
            cell.border = BORDER

        owner_cell = ws.cell(row=r, column=11)
        if owner in OWNER_FILLS:
            owner_cell.fill = OWNER_FILLS[owner]
            owner_cell.font = Font(bold=True)

        status_cell = ws.cell(row=r, column=12)
        status_cell.fill = STATUS_FILLS['Open']
        status_cell.font = Font(bold=True)

    widths = [16, 50, 8, 18, 16, 12, 16, 16, 8, 8, 16, 13, 30, 30, 35]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f'C{header_row + 1}'
    wb.save(OUTPUT)
    print(f'Saved: {OUTPUT}')
    print(f'Documents: {len(REGISTER)}')


if __name__ == '__main__':
    build()
