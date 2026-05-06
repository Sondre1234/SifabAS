"""SP-01415 Honeywell scope document list.

Filtered view of the SP-01415 SIF-1415-XXX document register, showing only
documents that Honeywell owns or contributes to. Used for the round with
Sidney Swart (Honeywell Enraf) so Honeywell has full control over what they
need to deliver, in the same SIF-XXXX numbering Hugin A used.

Saved to:
    05 Dokumentasjon / 03.Dok sendt leverandør for oppdatering /
        SP-01415_Document_List_Honeywell_Scope.xlsx
"""

import os
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from build_sp01415_tracker import REGISTER  # noqa: E402

from _paths import SHARED
PROJECT = SHARED / 'Zigma360' / 'Projects' / 'SP-01415 Small Volume Prover Snorre A'
DEST = PROJECT / '05 Dokumentasjon' / '03.Dok sendt leverandør for oppdatering'
DEST.mkdir(parents=True, exist_ok=True)
OUTPUT = DEST / 'SP-01415_Document_List_Honeywell_Scope.xlsx'

TITLE_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE_FONT = Font(bold=True, color='FFFFFF', size=14)
HEADER_FILL = PatternFill('solid', fgColor='2E75B6')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
SECTION_FILL = PatternFill('solid', fgColor='DDEBF7')
SECTION_FONT = Font(bold=True, color='1F4E78', size=11)
HW_FILL = PatternFill('solid', fgColor='F8CBAD')
HWSF_FILL = PatternFill('solid', fgColor='FFE699')
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='top')


# Filter register to rows where Honeywell has ownership or shared ownership.
HW_REGISTER = [
    row for row in REGISTER
    if row[4] in ('Honeywell', 'Sifab AS / HW')
]


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = 'HW Scope'

ws['A1'] = 'SP-01415 Snorre A SVP085 — Honeywell Scope Document List'
ws['A1'].font = TITLE_FONT
ws['A1'].fill = TITLE_FILL
ws.merge_cells('A1:K1')
ws.row_dimensions[1].height = 28

# Project meta
meta = [
    ('Project', 'SP-01415 Snorre A SVP085 (Honeywell SVP O085)'),
    ('Customer Project (Guidant)', 'GM-5341 Snorre A Oil Metering'),
    ('Honeywell Proposal', '10465986-O-1010834 R0'),
    ('Sifab PO (Guidant) received', '4500998501 dated 29 April 2026'),
    ('Honeywell PO (Sifab → Honeywell)', 'Pending — back-to-back to follow'),
    ('Reference project', 'Hugin A SP-00525 (SVP050) document delivery package'),
    ('Document', 'Honeywell-scope document register (transmittal-level, SIF-1415-XXX)'),
    ('Prepared by', 'Sifab AS — Sondre Falch'),
    ('Revision', 'Working draft — 5 May 2026'),
]
for i, (k, v) in enumerate(meta):
    r = 2 + i
    ws.cell(row=r, column=1, value=k).font = Font(bold=True)
    ws.cell(row=r, column=2, value=v)
    ws.cell(row=r, column=1).alignment = WRAP
    ws.cell(row=r, column=2).alignment = WRAP

note_row = 2 + len(meta) + 1
ws.cell(row=note_row, column=1, value='How to use').font = SECTION_FONT
ws.cell(row=note_row, column=1).fill = SECTION_FILL
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=11)

instructions = [
    'This list shows ONLY documents where Honeywell has ownership or shared ownership.',
    'Honeywell to fill in: "In HW Std Package? (Y/N/Partial)", "HW Internal Doc No", "HW Planned Delivery", "HW Comments".',
    'For shared scope (HW + Sifab), Honeywell fills the HW columns; Sifab handles the Sifab portion separately.',
    'Sifab Comments will be added before round 2.',
    'Document numbering (SIF-1415-XXX) mirrors Hugin A SP-00525 SIF-525-XXX where applicable — see Notes column.',
    'The ITP / Inspection & Test Plan, modular split drawings + lifting per NORSOK R-002, NORSOK datasheets, '
    'and SAT procedure are all SHARED scope and need Honeywell input even if Sifab leads.',
]
for i, txt in enumerate(instructions):
    ws.cell(row=note_row + 1 + i, column=1, value=txt).alignment = WRAP
    ws.merge_cells(
        start_row=note_row + 1 + i,
        start_column=1,
        end_row=note_row + 1 + i,
        end_column=11,
    )

# Column header
hdr_row = note_row + len(instructions) + 2
columns = [
    'Document No.',
    'Document Title',
    'Submit/Retain (S/R)',
    'Sifab Planned (WAO/WD)',
    'Owner',
    'In HW Std Package? (Y/N/Partial)',
    'HW Internal Doc No',
    'HW Planned Delivery',
    'HW Comments',
    'Sifab Comments',
    'Notes / Sifab cross-ref',
]
for c_idx, name in enumerate(columns, start=1):
    cell = ws.cell(row=hdr_row, column=c_idx, value=name)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = HEADER_ALIGN
    cell.border = BORDER
ws.row_dimensions[hdr_row].height = 36

# Data rows
data_row = hdr_row + 1
for sif_id, title, sr, due, owner, notes in HW_REGISTER:
    values = [sif_id, title, sr, due, owner, '', '', '', '', '', notes]
    for c_idx, val in enumerate(values, start=1):
        c = ws.cell(row=data_row, column=c_idx, value=val)
        c.alignment = WRAP
        c.border = BORDER
    # Owner colour (column 5)
    own_cell = ws.cell(row=data_row, column=5)
    if owner == 'Honeywell':
        own_cell.fill = HW_FILL
    elif owner == 'Sifab AS / HW':
        own_cell.fill = HWSF_FILL
    own_cell.font = Font(bold=True)
    data_row += 1

# Column widths
widths = [16, 50, 8, 18, 16, 16, 18, 18, 32, 32, 35]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = f'C{hdr_row + 1}'

wb.save(OUTPUT)
print(f'Saved: {OUTPUT}')
print(f'Honeywell-scope rows: {len(HW_REGISTER)} (of {len(REGISTER)} total)')
