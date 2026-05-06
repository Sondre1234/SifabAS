"""SP-01415 Procurement Tracker.

One row per Sifab purchase or sub-contract for the Snorre A SVP project.
Includes free-issue items Sifab supplies to suppliers (6Mo tubing, BFOU
cables, RTDs, thermowells, B&B valve, transmitters, edge protection, etc.),
direct supplier purchases (Honeywell, Pemberton, Flowtec, paint shop),
sub-contracts (Blu Electro, Intertek, Håkull) and banking instruments.

Owned by the procurement agent. Auto-scan via tools/sp01415_scan_procurement.py
(planned — uses same pattern as document scanner).

Saved at project root:
    Zigma360 / Projects / SP-01415 .../ SP-01415_Procurement_Tracker.xlsx
"""

import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from _paths import SHARED
PROJECT = SHARED / 'Zigma360' / 'Projects' / 'SP-01415 Small Volume Prover Snorre A'
OUTPUT = PROJECT / 'SP-01415_Procurement_Tracker.xlsx'

TITLE_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE_FONT = Font(bold=True, color='FFFFFF', size=14)
HEADER_FILL = PatternFill('solid', fgColor='2E75B6')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
SECTION_FILL = PatternFill('solid', fgColor='DDEBF7')
SECTION_FONT = Font(bold=True, color='1F4E78', size=11)

CATEGORY_FILLS = {
    'Direct supply': PatternFill('solid', fgColor='F8CBAD'),
    'Free-issue': PatternFill('solid', fgColor='C6E0B4'),
    'Sub-contract': PatternFill('solid', fgColor='BDD7EE'),
    'Service': PatternFill('solid', fgColor='FFE699'),
    'Banking': PatternFill('solid', fgColor='D9D9D9'),
}

STATUS_FILLS = {
    'Open': PatternFill('solid', fgColor='F4B084'),
    'RFQ issued': PatternFill('solid', fgColor='FFD966'),
    'Quoted': PatternFill('solid', fgColor='FFE699'),
    'PO issued': PatternFill('solid', fgColor='9DC3E6'),
    'In production': PatternFill('solid', fgColor='9DC3E6'),
    'In transit': PatternFill('solid', fgColor='FFD966'),
    'Received': PatternFill('solid', fgColor='A9D08E'),
    'Closed': PatternFill('solid', fgColor='A9D08E'),
}

THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='top')


# ---------------------------------------------------------------------------
# Procurement register
# (id, category, supplier, scope, qty, currency, est_cost, lead_time,
#  need_by, status, notes)
# Cost in original currency (USD or NOK). Convert in Financial Overview.
# ---------------------------------------------------------------------------
SECTIONS = [
    ('Direct supply — primary', [
        ('PR-1415-01', 'Direct supply', 'Honeywell BV (Schiphol, NL)',
         'SVP Snorre A — Standard prover (SVP085-O-E8C06N5A-2A-000-112-211-YYY-01). '
         'Sifab PO-00731 Pos 1, dated 2026-05-06. Ref Honeywell proposal '
         '10465986-O-1010834 R2, CTL-102003.', 1, 'USD',
         576319.53, '42 weeks',
         '~2027-02-15', 'PO issued',
         'PO-00731 issued via Zigma to Honeywell BV (Schiphol). EXW, net 30 days. Contact: '
         'Sidney Swart - Orders. Awaiting Honeywell order acknowledgement.'),
        ('PR-1415-02', 'Direct supply', 'Honeywell BV (Schiphol, NL)',
         'Quoted spare parts package (3 of 4 items: KIT OPTICAL SWITCH set of 3 '
         'P/N 44110148, MOTOR STOP SWITCH O P/N 44107515, SEAL KIT CSK S85 150/300/600# '
         'CRB P/N 44105627). RELAY 3 PHASE NOT ordered — note on PO.', 1, 'USD',
         14222.07, 'with prover',
         '~2027-02-15', 'PO issued',
         'PO-00731 Pos 2. Sum at list price (no 5% CP discount applied). '
         '14,222.07 ≈ 1,613 + 1,611 + 10,998. RELAY 3 PHASE 530V deliberately excluded.'),
        ('PR-1415-04', 'Direct supply', 'Pemberton (US)',
         'Seraphin Can EMSS0283.9L-30-3, 283.9 L, 316 SS, NIST + NVLAP cal — quote Pos 2',
         1, 'USD', 75054.92, '8–12 weeks',
         'Pre-FAT (~2026-12)', 'Open',
         'SS316 cabinet (not aluminum as in old quote). Pemberton NIST + NVLAP cert accepted (no Justervesenet — precedent NOA/Valhall).'),
        ('PR-1415-05', 'Direct supply', 'Flowtec',
         'Water draw kit parts (solenoid valve + manual instrument valves)', 1, 'NOK',
         200000, '4–6 weeks',
         '~2026-09-30', 'Open',
         'RFQ already issued. Confirm against final design (1\" tubing per lessons learned).'),
    ]),

    ('Free-issue items — Sifab supplies to Honeywell / installs in Norway', [
        ('FI-1415-01', 'Free-issue', 'TBD (Norwegian instrument supplier)',
         '3× pressure transmitters (Sifab free-issue per RFQ §1.6-8). 316SS body. ATEX. '
         'Tag IDs from Guidant.', 3, 'NOK',
         180000, '6–8 weeks',
         'Pre-fab kit (~2026-08)', 'Open',
         'Lesson C-PM-03: tag IDs needed early. Pre-manufacture kit and ship to Honeywell.'),
        ('FI-1415-02', 'Free-issue', 'TBD',
         'B&B (DBB) valve ½″ per VDS-MHBD102R, TR2000 BD20X — 1× pressure take-off',
         1, 'NOK', 25000, '4–6 weeks',
         '~2026-08-15', 'Open',
         'Sifab scope per RFQ §1.6-7. NACE compliance required.'),
        ('FI-1415-03', 'Free-issue', 'Guidant (free-issue to Sifab) → installed by Sifab',
         '3× process RTD sensors + 1× rod RTD sensor — Guidant free-issues; Sifab installs',
         4, 'NOK', 0, 'TBD by Guidant',
         '~2026-08-15', 'Open',
         'Awaiting Guidant supply schedule. Lesson C-PM-02: get firm need-dates.'),
        ('FI-1415-04', 'Free-issue', 'TBD (Sifab procurement)',
         'Thermowells: 3× process + 1× rod, ½″ NPTF, bore 6.5 mm, TR2000 BD20X material',
         4, 'NOK', 60000, '4–6 weeks',
         '~2026-08-15', 'Open',
         'Wake frequency calcs required (RFQ §1.16-9). Material per BD20X.'),
        ('FI-1415-05', 'Free-issue', 'TBD (Sifab procurement)',
         '6Mo tubing + Hoke Gyrolok metric fittings per ST701/SF712, all instrument hookups',
         1, 'NOK', 80000, '4–6 weeks',
         '~2026-08-15', 'Open',
         'Aker BP precedent: Sifab supplies 6Mo as free-issue. DEV-009.'),
        ('FI-1415-06', 'Free-issue', 'TBD',
         'BFOU halogen-free cables (motor, optical switches, JB to controller)',
         1, 'NOK', 50000, '2–4 weeks',
         '~2026-08-15', 'Open',
         'DEV-011. Per NORSOK E-001 / TR3023.'),
        ('FI-1415-07', 'Free-issue', 'TBD',
         'Cable glands SS316 / nickel-plated brass per E-001 / TR3023', 1, 'NOK',
         15000, '2–3 weeks',
         '~2026-08-15', 'Open',
         'DEV-011. Lesson C-EI-05: clear spec to factory.'),
        ('FI-1415-08', 'Free-issue', 'TBD',
         '6Mo Parker plugs for vent valves (lesson C-EI-01)', 1, 'NOK',
         5000, '2 weeks',
         '~2026-08-15', 'Open',
         'Lessons learned NOA punch #11: avoid carbon-steel plugs.'),
        ('FI-1415-09', 'Free-issue', 'TBD',
         'SS316 cable tray edge protection ("Volvo list")', 1, 'NOK',
         5000, '2 weeks',
         '~2026-08-15', 'Open',
         'Lesson C-EI-05. Generic spec + photos with Factory Instruction Package.'),
        ('FI-1415-10', 'Free-issue', 'TBD',
         '100 mm insulation material for flow tube + nozzles (post-NORSOK paint)',
         1, 'NOK', 25000, '2–3 weeks',
         '~2027-02-01', 'Open',
         'DEV-015. Apply in Norway after paint.'),
    ]),

    ('NORSOK painting (Sifab in-Norway scope)', [
        ('SC-1415-01', 'Sub-contract', 'TBD (Norwegian NORSOK paint shop)',
         'NORSOK M-501 System 6C coating of flow tube + wetted parts. '
         'Polyurethane-free per Equinor TR0042. Operator qualification required.',
         1, 'NOK', 130000, '2–3 weeks',
         '~2027-01-15', 'Open',
         'DEV-004. Identify and qualify shop. Submit procedure to Guidant within 2 WAO.'),
    ]),

    ('Sub-contracts — onshore + offshore execution', [
        ('SC-1415-02', 'Sub-contract', 'Honeywell Enraf supervisor',
         'Onshore disassembly supervision at Sifab/Norwegian fab yard. '
         'Approx 120 hours @ USD 140 + KPI', 120, 'USD',
         16800, 'On dispatch',
         'Q1 2027', 'Open',
         'Quote Pos 4. Pos 4 lump sum ~USD 110k expected. VOR per PP-PS-13.'),
        ('SC-1415-03', 'Sub-contract', 'Honeywell Enraf supervisor',
         'Offshore re-assembly supervision on Snorre A. Approx 180–240 hours '
         '@ USD 180 + KPI + mob/demob', 240, 'USD',
         43200, 'On dispatch',
         'Q1–Q2 2027', 'Open',
         'Quote Pos 4. KPI adjustment per project. Travel from US/Bosnia.'),
        ('SC-1415-04', 'Sub-contract', 'Intertek (prover expert)',
         'Onshore + offshore prover expert support alongside HE supervisor. '
         '~36 h onshore @ $140 + ~180–240 h offshore @ $180 (+ KPI)',
         276, 'USD', 38640, 'On dispatch',
         'Q1–Q2 2027', 'Open',
         'Quote Pos 4. Engages with HE supervisor on disassembly + offshore re-assembly.'),
        ('SC-1415-05', 'Sub-contract', 'Blu Electro',
         'Offshore cabling re-connect: cables, JBs, cable trays parallel with re-assembly',
         1, 'NOK', 200000, 'On dispatch',
         'Q1–Q2 2027', 'Open',
         'Quote Pos 4. Sifab cabling partner.'),
    ]),

    ('Logistics + freight forwarding', [
        ('SV-1415-01', 'Service', 'Air freight (Honeywell → Norway)',
         'Air freight prover skid in wooden crate, 10 t lifting capability. '
         'TruStop AZ → Sifab Sandnes', 1, 'USD',
         58798, '~5–10 days',
         '~2027-02-15', 'Open',
         'Quote Pos 3. Sea freight cheaper alternative if schedule allows.'),
        ('SV-1415-02', 'Service', 'Håkull AS (Sandnes — Even Hansen)',
         'Onshore transport Sifab → quayside for offshore lift to Snorre A', 1, 'NOK',
         50000, '1–2 weeks',
         '~2027-03-15', 'Open',
         'Local freight forwarder. Norwegian-language correspondence (memory).'),
        ('SV-1415-03', 'Service', 'Travel — Tempe FAT',
         'Travel + accommodation Tempe, Arizona — kick-off + FAT + calibration',
         1, 'USD', 30104.74, 'Per visit',
         '~2026-12 / 2027-01', 'Open',
         'Quote Pos 6. Sifab arranges trips.'),
    ]),

    ('Banking + financial instruments', [
        ('BG-1415-01', 'Banking', 'Sifab bank (TBD)',
         'Bank guarantee 1-year valid, beneficiary Measurement Solutions Norway AS, '
         'amount USD 482,479 (= Guidant M1 30%)', 1, 'USD',
         3000, '1–2 weeks',
         '~2026-05-20', 'Open',
         'M1 condition. Sondre confirmed 1-yr to Anne 2026-04-30. Fee estimate ~0.5–1%/yr on USD value.'),
        ('BG-1415-02', 'Banking', 'Sifab bank (TBD)',
         'Credit facility for project cashflow (Honeywell down vs Guidant M1 gap)',
         1, 'NOK', 2000, '2–3 weeks',
         '~2026-05-20', 'Open',
         'Lesson C-FIN-03: project >NOK 15M. Arrange before any supplier down-payment exposure.'),
    ]),
]


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------
def build():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Procurement'

    ws['A1'] = 'SP-01415 Snorre A SVP085 — Procurement Tracker'
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = TITLE_FILL
    ws.merge_cells('A1:N1')
    ws.row_dimensions[1].height = 28

    meta = [
        ('Project', 'SP-01415 Snorre A SVP085'),
        ('Customer Project', 'GM-5341 Snorre A Oil Metering'),
        ('Customer PO', '4500998501 (USD 1,608,262.36)'),
        ('Owner agent', 'procurement (with commercial-finance for cost roll-up)'),
        ('Last updated', '2026-05-06'),
    ]
    for i, (k, v) in enumerate(meta):
        r = 2 + i
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
        ws.cell(row=r, column=1).alignment = WRAP
        ws.cell(row=r, column=2).alignment = WRAP

    note_row = 2 + len(meta) + 1
    ws.cell(row=note_row, column=1, value='How to use').font = SECTION_FONT
    ws.cell(row=note_row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=14)

    instructions = [
        '1. One row per Sifab purchase: direct supply, free-issue procurement, sub-contract, service, banking.',
        '2. Costs in original currency (USD or NOK). Conversion done in Financial Overview.',
        '3. Need-by dates work backwards from the customer PO delivery 2027-04-15.',
        '4. Status flow: Open → RFQ issued → Quoted → PO issued → In production / In transit → Received → Closed.',
        '5. Free-issue rows: Sifab procures and ships kit to Honeywell or installs in Norway.',
        '6. Honeywell back-to-back PO MUST have 28 mo warranty from platform install (matches Guidant PO).',
        '7. Bank guarantee for Guidant M1 must be in place before any Honeywell down-payment.',
    ]
    for i, txt in enumerate(instructions):
        ws.cell(row=note_row + 1 + i, column=1, value=txt).alignment = WRAP
        ws.merge_cells(
            start_row=note_row + 1 + i, start_column=1,
            end_row=note_row + 1 + i, end_column=14,
        )

    hdr_row = note_row + len(instructions) + 2
    columns = [
        'ID',
        'Category',
        'Supplier',
        'Scope',
        'Qty',
        'Currency',
        'Estimated cost',
        'Lead time',
        'Need-by date',
        'PO No (Sifab)',
        'PO date',
        'Status',
        'Filename / folder',
        'Notes',
    ]
    for c_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=hdr_row, column=c_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    ws.row_dimensions[hdr_row].height = 36

    current_row = hdr_row + 1
    grand_total_usd = 0.0
    grand_total_nok = 0.0

    for section_title, items in SECTIONS:
        cell = ws.cell(row=current_row, column=1, value=section_title)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.alignment = Alignment(vertical='center')
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=14)
        current_row += 1

        for (pid, cat, supplier, scope, qty, cur, cost, lead, need_by, status, notes) in items:
            values = [pid, cat, supplier, scope, qty, cur, cost, lead,
                      need_by, '', '', status, '', notes]
            for c_idx, val in enumerate(values, start=1):
                c = ws.cell(row=current_row, column=c_idx, value=val)
                c.alignment = WRAP
                c.border = BORDER

            cat_cell = ws.cell(row=current_row, column=2)
            if cat in CATEGORY_FILLS:
                cat_cell.fill = CATEGORY_FILLS[cat]
                cat_cell.font = Font(bold=True)

            status_cell = ws.cell(row=current_row, column=12)
            if status in STATUS_FILLS:
                status_cell.fill = STATUS_FILLS[status]
                status_cell.font = Font(bold=True)

            cost_cell = ws.cell(row=current_row, column=7)
            cost_cell.number_format = '#,##0.00'

            if cur == 'USD':
                grand_total_usd += cost or 0
            elif cur == 'NOK':
                grand_total_nok += cost or 0

            current_row += 1

    # Totals
    current_row += 1
    ws.cell(row=current_row, column=3, value='Total estimated cost (USD)').font = Font(bold=True)
    ws.cell(row=current_row, column=6, value='USD').font = Font(bold=True)
    c = ws.cell(row=current_row, column=7, value=grand_total_usd)
    c.font = Font(bold=True)
    c.number_format = '#,##0.00'
    current_row += 1
    ws.cell(row=current_row, column=3, value='Total estimated cost (NOK)').font = Font(bold=True)
    ws.cell(row=current_row, column=6, value='NOK').font = Font(bold=True)
    c = ws.cell(row=current_row, column=7, value=grand_total_nok)
    c.font = Font(bold=True)
    c.number_format = '#,##0.00'

    widths = [12, 14, 30, 60, 6, 9, 14, 12, 16, 14, 12, 14, 28, 35]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f'C{hdr_row + 1}'
    wb.save(OUTPUT)
    total = sum(len(items) for _, items in SECTIONS)
    print(f'Saved: {OUTPUT}')
    print(f'Sections: {len(SECTIONS)} | Lines: {total}')
    print(f'Total estimated cost: USD {grand_total_usd:,.2f} + NOK {grand_total_nok:,.2f}')


if __name__ == '__main__':
    build()
