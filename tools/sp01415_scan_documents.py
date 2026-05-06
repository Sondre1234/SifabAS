"""SP-01415 Snorre A — manual document scan.

Run on user request only ("vi scanner etter dokumenter").

WORKFLOW NOTE — important
-------------------------
This scan only sees the mailbox of the Microsoft 365 account it is signed in
with (Sondre's). Emails Tom (or anyone else) sends DIRECTLY to Honeywell or
Guidant without copying Sondre will NOT appear here.

Rule of thumb: anyone in Sifab who sends a project email to a supplier or
customer should CC sondre.falch@sifab.no (or post@sifab.no if a shared
mailbox is used) so the agent scan can pick the email + attachments up.

What it does:
  1. Searches inbox/sent/deleted via two parallel passes:
        a) Project keyword pass — Snorre, SP-01415, PO 4500998501, GM-5341,
           GM-8501-1447, etc. Catches anything that mentions the project.
        b) Domain pass — any email from/to @honeywell.com, @guidantmeasurement.com,
           and other supplier/customer domains. Catches Tom's direct
           correspondence even when keywords are missing, provided Sondre was
           on To/CC/BCC.
     Results are merged.
  2. Optionally downloads attachments not already filed (--download).
  3. Scans the project subfolders on OneDrive and lists all files found
     with their modified date.
  4. Matches files against the document register (SIF-1415-XXX) by fuzzy
     keyword match against the document title.
  5. Updates the tracker workbook in-place:
        - "Filename(s) found in folders"
        - "Folder location"
        - "Date for latest Submission" (uses file mtime as proxy)
        - "Status" → "Draft" if file found locally only,
                     "Submitted" if email evidence of customer send is found.
  6. Prints a summary report.

Read-only by default. Pass --download to actually save email attachments to
the suggested project subfolder. Pass --update to update the tracker in place.

Usage:
    python tools/sp01415_scan_documents.py                # report only
    python tools/sp01415_scan_documents.py --update       # also write tracker
    python tools/sp01415_scan_documents.py --download     # also save attachments
    python tools/sp01415_scan_documents.py --update --download
"""

from __future__ import annotations

import argparse
import base64
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

# Make project tools importable
sys.path.insert(0, str(Path(__file__).parent))

from openpyxl import load_workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill  # noqa: E402

# We reuse the email_client auth + helpers
from email_client import api_get, get_token  # noqa: E402

from _paths import SHARED
PROJECT = SHARED / 'Zigma360' / 'Projects' / 'SP-01415 Small Volume Prover Snorre A'
TRACKER = PROJECT / 'SP-01415_Dokument_status_og_oppfølging.xlsx'

# Keywords that mark an email as project-relevant
SEARCH_KEYWORDS = [
    'SP-01415',
    'PO 4500998501',
    'PO-00731',
    'GM-5341',
    'GM-8501-1447',
    'Snorre A',
    'Small Volume Prover Snorre',
    'SVP085 Snorre',
    'O-1010834',  # Honeywell proposal reference
    'CTL-102003',  # Honeywell CTL discount reference
]

# Domains that, when present in any sender or recipient address, mark the
# email as project-relevant even if no keyword is matched. Catches direct
# Sifab ↔ supplier/customer correspondence (e.g. when Tom sends to Sidney
# without using a project keyword in the subject) PROVIDED Sondre is on the
# To/CC/BCC line — only Sondre's mailbox is visible to this scan.
PROJECT_DOMAINS = [
    'guidantmeasurement.com',
    'honeywell.com',
    'pemberton',                # Seraphin can supplier
    'flowtec',                  # Water draw kit parts
    'haakull',                  # Freight forwarder
    'intertek',                 # Offshore prover expert
    'bluelectro',               # Cabling subcontractor
]

# Only consider emails after the project started (RFQ first received 2026-02-23).
PROJECT_START = '2026-02-01'

# Skip these attachment names (inline-signature images, generic broschures).
NOISE_ATT_PATTERNS = [
    r'^image\d{3}\.(jpg|png|gif|jpeg)$',
    r'^Skids and Special items\.pdf$',
    r'^Flow metering and instrumentation broshure\.pdf$',
    r'^Sifab AS Company Profile.*\.pdf$',
    r'^Workload SIFAB.*\.xlsx$',
    r'^V-Cone Norway.*\.xlsx$',
    r'^Medlemsliste.*\.xlsx$',
]
# Minimum attachment size in bytes (filter out tiny inline images).
MIN_ATT_SIZE = 30 * 1024  # 30 KB

# Folders we scan for already-filed documents (relative to PROJECT).
# 01.Dokumentmaler is intentionally EXCLUDED — those are blank templates,
# not submitted documents.
SCAN_SUBFOLDERS = [
    '01.Bestillinger & Ordrebekreftelser',
    '02 Tilbud',
    '03 Underlag fra Kunde',
    '04 E-mail',
    '05 Dokumentasjon/02.Dok fra Leverandør',
    '05 Dokumentasjon/03.Dok sendt leverandør for oppdatering',
    '05 Dokumentasjon/04.Dok sendt til kunde',
    '05 Dokumentasjon/05.Dok Kode 1',
    '05 Dokumentasjon/Transmittel',
    '06 Møtereferat',
    '07 Clarifications',
    '08 Underlag fra Leverandører',
    '09 VOR',
    '11 Punch lists',
]

# Files matching any of these patterns (case-insensitive) are not counted as
# project submissions. They are templates, RFQs, or internal tooling that
# happen to live in the project tree.
EXCLUDE_FILE_PATTERNS = [
    r'template',
    r'\bv-?cone\b',
    r'^rfq[ _]',                  # RFQ documents (we send these, not deliverables)
    r'_rfq[ _]',
    r'document.list.for.honeywell',  # internal Honeywell-review list
    r'broshure|broschure',
    r'^~\$',
]

# Files older than this are not counted as SP-01415 submissions.
MIN_FILE_DATE = date(2026, 2, 1)

# Fuzzy keyword index used to match a filename against a SIF-1415-XXX doc.
# Each entry: (SIF id, list of regex patterns the filename should match)
KEYWORDS_PER_DOC = {
    'SIF-1415-001': [r'sdl', r'document.list', r'supplier.document'],
    'SIF-1415-002': [r'\bitp\b', r'\bqip\b', r'inspection.test.plan',
                     r'quality.inspection.plan'],
    'SIF-1415-003': [r'progress.report', r'monthly.progress'],
    'SIF-1415-004': [r'norsok.datasheet', r'norsok.instrument'],
    'SIF-1415-005': [r'sizing.calculation', r'design.calculation',
                     r'sizing.report'],
    'SIF-1415-006': [r'wake.frequency'],
    'SIF-1415-007': [r'ga.drawing.prover', r'general.arrangement.prover'],
    'SIF-1415-008': [r'ga.*thermowell', r'thermowell.drawing'],
    'SIF-1415-009': [r'seraphin.*can', r'pemberton', r'cabinet.*seraphin'],
    'SIF-1415-010': [r'water.draw.panel', r'flowtec'],
    'SIF-1415-011': [r'wiring.diagram', r'product.datasheet'],
    'SIF-1415-012': [r'split.module', r'modular.split', r'lifting.plan'],
    'SIF-1415-013': [r'3d.model', r'\.stp$', r'step.model'],
    'SIF-1415-014': [r'fat.procedure', r'calibration.procedure'],
    'SIF-1415-015': [r'hydrostatic', r'pressure.test.procedure'],
    'SIF-1415-016': [r'pmi.procedure'],
    'SIF-1415-017': [r'\bndt.procedure\b', r'\bnde.procedure\b'],
    'SIF-1415-018': [r'ndt.operator', r'nde.operator'],
    'SIF-1415-019': [r'surface.treatment.procedure', r'paint.procedure',
                     r'coating.procedure'],
    'SIF-1415-020': [r'\bwps\b', r'\bwpqr\b', r'welder.cert',
                     r'welding.document'],
    'SIF-1415-021': [r'disassembly.procedure', r'packing.procedure'],
    'SIF-1415-022': [r're.assembly.procedure', r'reassembly.procedure',
                     r'offshore.assembly'],
    'SIF-1415-023': [r'sat.procedure', r'water.draw.procedure'],
    'SIF-1415-024': [r'lifting.plan', r'r.002', r'lifting.procedure'],
    'SIF-1415-025': [r'preservation.procedure'],
    'SIF-1415-026': [r'\bspir\b'],
    'SIF-1415-027': [r'material.cert', r'mtr', r'10204'],
    'SIF-1415-028': [r'pressure.test.report'],
    'SIF-1415-029': [r'ped.cert', r'ec.declaration', r'declaration.of.conformity'],
    'SIF-1415-030': [r'fat.report', r'calibration.report'],
    'SIF-1415-031': [r'pmi.report'],
    'SIF-1415-032': [r'\biom\b', r'operation.maintenance.manual',
                     r'service.manual'],
    'SIF-1415-033': [r'ndt.report', r'nde.report'],
    'SIF-1415-034': [r'surface.treatment.report', r'paint.report',
                     r'coating.report'],
    'SIF-1415-035': [r'weld.report', r'weld.record'],
    'SIF-1415-036': [r'megger', r'meggering'],
    'SIF-1415-037': [r'motor.document', r'motor.spec'],
    'SIF-1415-038': [r'sat.report', r'water.draw.report'],
    'SIF-1415-039': [r'weighing.report', r'weight.report'],
    'SIF-1415-040': [r'preservation.report'],
    'SIF-1415-041': [r'justervesenet'],
    'SIF-1415-042': [r'atex.cert', r'atex.certificate', r'svp.atex'],
    'SIF-1415-043': [r'pre.fat.punch'],
    'SIF-1415-044': [r'fat.punch'],
    'SIF-1415-045': [r'sat.punch'],
    'SIF-1415-046': [r'milestone.completion', r'milestone.cert'],
    'SIF-1415-047': [r'sub.supplied.items', r'sub.supplier.status'],
    'SIF-1415-048': [r'deviation.list'],
    'SIF-1415-049': [r'kick.off', r'mom', r'meeting.minutes'],
    'SIF-1415-050': [r'close.out.report', r'closeout'],
}

# Where attachments from a sender should be filed by default.
SENDER_FILE_MAP = [
    # (substring of address, sub-folder)
    ('guidantmeasurement.com',
     '03 Underlag fra Kunde'),                           # client → customer-provided
    ('honeywell',
     '05 Dokumentasjon/02.Dok fra Leverandør'),          # supplier → docs from supplier
    ('hultec',
     '05 Dokumentasjon/02.Dok fra Leverandør'),
    ('pemberton',
     '08 Underlag fra Leverandører'),
    ('flowtec',
     '08 Underlag fra Leverandører'),
    ('haakull',
     '08 Underlag fra Leverandører'),
    ('intertek',
     '08 Underlag fra Leverandører'),
]
DEFAULT_FILE_DEST = '04 E-mail'  # fallback


# ---------------------------------------------------------------------------
# Email scan
# ---------------------------------------------------------------------------
def _strip_html(html: str) -> str:
    return re.sub(r'<[^>]+>', '', html or '').strip()


def search_emails(token: str):
    """Return list of dicts with id, subject, from, received, hasAttachments,
    matchedBy. Two parallel passes:
        keyword — full-text $search across folders for project terms;
        domain  — $search for each project-relevant domain.
    Results merged on id.
    """
    seen: dict[str, dict] = {}
    folders = ['inbox', 'sentitems', 'deleteditems']

    def store(m: dict, folder: str, matched_by: str):
        mid = m['id']
        received = m.get('receivedDateTime', '')
        if received and received[:10] < PROJECT_START:
            return
        if mid in seen:
            # Append match reason for visibility
            tags = set(seen[mid]['matchedBy'].split(','))
            tags.add(matched_by)
            seen[mid]['matchedBy'] = ','.join(sorted(tags))
            return
        seen[mid] = {
            'id': mid,
            'folder': folder,
            'subject': m.get('subject', ''),
            'from': (m.get('from') or {}).get('emailAddress', {}).get('address', ''),
            'received': received,
            'hasAttachments': m.get('hasAttachments', False),
            'matchedBy': matched_by,
        }

    # Pass 1 — keyword search
    for folder in folders:
        for kw in SEARCH_KEYWORDS:
            endpoint = f'/me/mailFolders/{folder}/messages?$search="{kw}"&$top=50'
            try:
                data = api_get(endpoint, token)
            except Exception as exc:
                print(f'  ! keyword search failed in {folder} for "{kw}": {exc}')
                continue
            for m in data.get('value', []):
                store(m, folder, f'kw:{kw}')

    # Pass 2 — domain search (catches Tom's direct correspondence)
    for folder in folders:
        for dom in PROJECT_DOMAINS:
            endpoint = f'/me/mailFolders/{folder}/messages?$search="{dom}"&$top=50'
            try:
                data = api_get(endpoint, token)
            except Exception as exc:
                print(f'  ! domain search failed in {folder} for "{dom}": {exc}')
                continue
            for m in data.get('value', []):
                store(m, folder, f'dom:{dom}')

    return sorted(seen.values(), key=lambda x: x['received'])


def list_attachments(token: str, msg_id: str):
    """Return list of attachment dicts: name, contentType, size, contentBytes (b64).
    Filters out inline-signature noise and tiny images.
    """
    data = api_get(f'/me/messages/{msg_id}/attachments', token)
    out = []
    for att in data.get('value', []):
        if att.get('@odata.type') != '#microsoft.graph.fileAttachment':
            continue
        if att.get('isInline'):
            continue
        name = att.get('name', 'unnamed')
        size = att.get('size', 0)
        if size < MIN_ATT_SIZE:
            continue
        if any(re.match(p, name, re.I) for p in NOISE_ATT_PATTERNS):
            continue
        out.append({
            'name': name,
            'contentType': att.get('contentType', ''),
            'size': size,
            'contentBytes': att.get('contentBytes'),
        })
    return out


def file_destination_for_sender(sender: str) -> str:
    sender_lc = (sender or '').lower()
    for needle, dest in SENDER_FILE_MAP:
        if needle in sender_lc:
            return dest
    return DEFAULT_FILE_DEST


def save_attachment(att: dict, dest_folder: Path) -> Path:
    dest_folder.mkdir(parents=True, exist_ok=True)
    safe_name = re.sub(r'[<>:"/\\|?*]', '_', att['name'])
    out = dest_folder / safe_name
    if out.exists():
        return out  # don't overwrite
    out.write_bytes(base64.b64decode(att['contentBytes']))
    return out


# ---------------------------------------------------------------------------
# Folder scan
# ---------------------------------------------------------------------------
def scan_folders():
    """Return list of dicts: path, name, folder, mtime — excluding templates,
    RFQs, V-Cone files, and anything older than MIN_FILE_DATE."""
    files = []
    excluded_count = 0
    for sub in SCAN_SUBFOLDERS:
        full = PROJECT / sub
        if not full.exists():
            continue
        for f in full.rglob('*'):
            if not f.is_file() or f.name.startswith('~$'):
                continue
            # exclude by filename pattern
            if any(re.search(p, f.name, re.I) for p in EXCLUDE_FILE_PATTERNS):
                excluded_count += 1
                continue
            try:
                m = datetime.fromtimestamp(f.stat().st_mtime).date()
            except OSError:
                m = None
            # exclude old files (pre-project)
            if m and m < MIN_FILE_DATE:
                excluded_count += 1
                continue
            files.append({
                'path': f,
                'name': f.name,
                'folder': str(f.parent.relative_to(PROJECT)),
                'mtime': m,
            })
    if excluded_count:
        print(f'  ({excluded_count} files excluded as templates/RFQs/old/V-Cone)')
    return files


# ---------------------------------------------------------------------------
# Match files to register
# ---------------------------------------------------------------------------
def match_files_to_register(files):
    """Return dict: SIF-1415-XXX → list[file dict]"""
    matches = {sif: [] for sif in KEYWORDS_PER_DOC}

    for f in files:
        haystack = f['name'].lower().replace('_', ' ').replace('-', ' ')
        for sif, patterns in KEYWORDS_PER_DOC.items():
            for pat in patterns:
                if re.search(pat, haystack):
                    matches[sif].append(f)
                    break
    return matches


# ---------------------------------------------------------------------------
# Update tracker
# ---------------------------------------------------------------------------
DATE_FMT = '%Y-%m-%d'
STATUS_FILLS = {
    'Open': PatternFill('solid', fgColor='F4B084'),
    'Draft': PatternFill('solid', fgColor='FFE699'),
    'Submitted': PatternFill('solid', fgColor='9DC3E6'),
    'In review': PatternFill('solid', fgColor='FFD966'),
    'Approved': PatternFill('solid', fgColor='A9D08E'),
}

# Customer email-domain markers — a sent email with any recipient on this list
# counts as "Submitted to customer".
CUSTOMER_DOMAINS = ['guidantmeasurement.com']


def _is_customer_recipient(recipients: list[dict]) -> bool:
    for r in recipients or []:
        addr = (r.get('emailAddress') or {}).get('address', '').lower()
        if any(dom in addr for dom in CUSTOMER_DOMAINS):
            return True
    return False


def find_customer_submissions(token, project_emails):
    """For each project-relevant email, check sent items for attachments sent
    to a customer recipient. Returns dict {filename_lower: [(date, recipient)]}.
    """
    sent_attachments: dict[str, list[tuple[str, str]]] = {}

    # Look only at messages from sentitems folder among project-relevant emails
    sent_msgs = [e for e in project_emails if e['folder'] == 'sentitems']
    for em in sent_msgs:
        # Get full message with recipients
        try:
            data = api_get(
                f"/me/messages/{em['id']}?$select=subject,toRecipients,ccRecipients,sentDateTime,hasAttachments",
                token,
            )
        except Exception:
            continue
        recipients = (data.get('toRecipients') or []) + (data.get('ccRecipients') or [])
        if not _is_customer_recipient(recipients):
            continue
        if not data.get('hasAttachments'):
            continue

        atts = list_attachments(token, em['id'])
        sent_date = (data.get('sentDateTime') or em['received'])[:10]
        # Find first customer recipient address
        cust_addr = ''
        for r in recipients:
            a = (r.get('emailAddress') or {}).get('address', '').lower()
            if any(dom in a for dom in CUSTOMER_DOMAINS):
                cust_addr = a
                break
        for att in atts:
            key = att['name'].lower()
            sent_attachments.setdefault(key, []).append((sent_date, cust_addr))
    return sent_attachments


def update_tracker(matches, sent_to_customer=None):
    """Refresh status/dates/filenames for every doc.

    Rows with no matched files are reset to 'Open' so that false matches
    from a previous run are wiped clean."""
    if not TRACKER.exists():
        print(f'! Tracker not found: {TRACKER}')
        return

    wb = load_workbook(TRACKER)
    ws = wb['SDL']

    header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == 'Document No.':
            header_row = r
            break
    if header_row is None:
        print('! Could not locate header row in tracker')
        return

    sent_to_customer = sent_to_customer or {}
    submitted = drafted = reset = 0
    for r in range(header_row + 1, ws.max_row + 1):
        sif = ws.cell(row=r, column=1).value
        if not sif:
            continue
        files = sorted(
            [f for f in matches.get(sif, []) if f['mtime']],
            key=lambda f: f['mtime'],
            reverse=True,
        )

        status_cell = ws.cell(row=r, column=12)

        if not files:
            ws.cell(row=r, column=7, value='')
            ws.cell(row=r, column=9, value=0)
            ws.cell(row=r, column=13, value='')
            ws.cell(row=r, column=14, value='')
            status_cell.value = 'Open'
            status_cell.fill = STATUS_FILLS['Open']
            status_cell.font = Font(bold=True)
            reset += 1
            continue

        # Cross-reference filenames against sent-to-customer list
        sent_evidence = []   # list of (file dict, sent_date, recipient)
        for f in files:
            for sent_date, recipient in sent_to_customer.get(f['name'].lower(), []):
                sent_evidence.append((f, sent_date, recipient))

        names = '\n'.join(f['name'] for f in files[:5])
        if len(files) > 5:
            names += f'\n(+{len(files) - 5} more)'
        ws.cell(row=r, column=13, value=names)
        ws.cell(row=r, column=14, value='\n'.join(sorted({f['folder'] for f in files})))
        ws.cell(row=r, column=9, value=len(files))

        if sent_evidence:
            sent_evidence.sort(key=lambda t: t[1], reverse=True)
            latest_sent_date = sent_evidence[0][1]
            ws.cell(row=r, column=7, value=latest_sent_date)
            status_cell.value = 'Submitted'
            status_cell.fill = STATUS_FILLS['Submitted']
            status_cell.font = Font(bold=True)
            submitted += 1
        else:
            # Draft only — file exists locally but no email evidence of customer send
            ws.cell(row=r, column=7, value='')   # don't set submission date
            status_cell.value = 'Draft'
            status_cell.fill = STATUS_FILLS['Draft']
            status_cell.font = Font(bold=True)
            drafted += 1

    wb.save(TRACKER)
    print(
        f'Tracker updated: {submitted} submitted, {drafted} draft, '
        f'{reset} reset to Open.'
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument('--download', action='store_true',
                    help='Download new email attachments to project folders')
    ap.add_argument('--update', action='store_true',
                    help='Write findings into the tracker workbook')
    ap.add_argument('--limit-emails', type=int, default=20,
                    help='Max emails to inspect for attachments (default 20)')
    args = ap.parse_args()

    print('=' * 80)
    print(f'SP-01415 document scan — {date.today()}')
    print('=' * 80)

    # 1. Email scan
    print('\n[1/3] Email scan')
    print('-' * 80)
    token = get_token()
    emails = search_emails(token)
    print(f'  Found {len(emails)} project-relevant emails across inbox/sent/deleted')

    emails_with_att = [e for e in emails if e['hasAttachments']]
    print(f'  Of those, {len(emails_with_att)} have attachments')

    saved_attachments = 0
    for em in emails_with_att[: args.limit_emails]:
        attachments = list_attachments(token, em['id'])
        for att in attachments:
            dest_rel = file_destination_for_sender(em['from'])
            dest_dir = PROJECT / dest_rel
            target = dest_dir / re.sub(r'[<>:"/\\|?*]', '_', att['name'])
            if target.exists():
                continue
            print(
                f'  · {em["received"][:10]} | {em["from"]:<45} '
                f'→ {att["name"]} ({att["size"] // 1024} KB) → {dest_rel}'
            )
            if args.download:
                save_attachment(att, dest_dir)
                saved_attachments += 1
    if args.download:
        print(f'  Saved {saved_attachments} new attachments.')
    else:
        print('  (read-only — pass --download to save attachments)')

    # 2. Folder scan
    print('\n[2/3] Folder scan')
    print('-' * 80)
    files = scan_folders()
    print(f'  Files discovered in project subfolders: {len(files)}')

    # 3. Match + report
    print('\n[3/3] Match against document register (SIF-1415-XXX)')
    print('-' * 80)
    matches = match_files_to_register(files)
    matched = [(sif, lst) for sif, lst in matches.items() if lst]
    missing = [sif for sif, lst in matches.items() if not lst]

    print(f'  Matched: {len(matched)} / {len(KEYWORDS_PER_DOC)} documents have at least one file')
    print(f'  Missing: {len(missing)} documents (status will remain Open)')

    if matched:
        print('\n  Matched documents:')
        for sif, lst in matched:
            latest = max((f for f in lst if f['mtime']), key=lambda f: f['mtime'])
            print(f'    {sif}  ←  {latest["name"]:<70}  ({latest["mtime"]})')

    if missing:
        print('\n  Documents still missing:')
        for sif in missing:
            print(f'    {sif}')

    # 4. Find evidence of customer submissions in sent items
    print('\n[4/5] Customer-submission evidence (sent to @guidantmeasurement.com)')
    print('-' * 80)
    sent_to_customer = find_customer_submissions(token, emails)
    if sent_to_customer:
        for fname, evidence in sent_to_customer.items():
            for d, addr in evidence:
                print(f'  · {d}  {fname:<60}  → {addr}')
    else:
        print('  (none yet — no project document has been sent to Guidant)')

    # 5. Update tracker
    if args.update:
        print('\n[5/5] Update tracker')
        print('-' * 80)
        update_tracker(matches, sent_to_customer=sent_to_customer)

    print('\nDone.')


if __name__ == '__main__':
    main()
