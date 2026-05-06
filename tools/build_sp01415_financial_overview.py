"""SP-01415 Financial Overview (internal P&L).

Owned by the commercial-finance agent. Five tabs:
  1. Summary — headline P&L
  2. Revenue — milestones M1–M4 from Guidant PO
  3. Costs — direct supplier + free-issue + sub-contract + freight + banking
  4. Cashflow — month-by-month net position
  5. FX — USD/NOK exposure and hedge recommendation

Costs are pulled from the same numbers used in the Procurement Tracker so the
two stay in sync.

Saved at project root:
    Zigma360 / Projects / SP-01415 .../ SP-01415_Financial_Overview.xlsx
"""

import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHARED = Path(os.environ['USERPROFILE']) / 'OneDrive - Sifab AS' / 'Dokumenter - Felles'
PROJECT = SHARED / 'Zigma360' / 'Projects' / 'SP-01415 Small Volume Prover Snorre A'
OUTPUT = PROJECT / 'SP-01415_Financial_Overview.xlsx'

# FX assumption — update on hedge or at quote-vs-PO rate
USD_NOK = 10.65   # representative; commercial-finance to confirm and update

TITLE_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE_FONT = Font(bold=True, color='FFFFFF', size=14)
HEADER_FILL = PatternFill('solid', fgColor='2E75B6')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
SECTION_FILL = PatternFill('solid', fgColor='DDEBF7')
SECTION_FONT = Font(bold=True, color='1F4E78', size=11)

POS_FILL = PatternFill('solid', fgColor='A9D08E')
NEG_FILL = PatternFill('solid', fgColor='F8CBAD')

THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='top')

# ---------------------------------------------------------------------------
# Revenue (Guidant PO 4500998501)
# ---------------------------------------------------------------------------
PO_VALUE_USD = 1_608_262.36
REVENUE = [
    ('M1', 'Clarified PO + bank guarantee', 0.30, '~2026-05-15', 'Open',
     'BG to be issued. Honeywell 30% down out of M1 if back-to-back-aligned.'),
    ('M2', 'FAT accepted at Tempe AZ', 0.50, '~2026-12-15', 'Open',
     'Witnessed FAT pass.'),
    ('M3', 'Re-assembly + SAT complete on Snorre A', 0.10, '~2027-04-30', 'Open',
     'Sifab-controlled phase.'),
    ('M4', 'Documentation accepted, package close-out', 0.10, '~2027-08-31', 'Open',
     'Lesson C-PM-01: 6-month tail risk; escalate at 3 weeks.'),
]

# ---------------------------------------------------------------------------
# Costs — same numbers as Procurement Tracker
# ---------------------------------------------------------------------------
COSTS = [
    # Honeywell BV (Schiphol) — Sifab PO-00731 issued 2026-05-06
    ('Direct supply', 'Honeywell SVP Snorre A standard prover (Sifab PO-00731 Pos 1)',
     'USD', 576_319.53,
     'PO-00731 dated 2026-05-06 to Honeywell BV. Ref proposal R2, CTL-102003.'),
    ('Direct supply', 'Honeywell quoted spare parts package (Sifab PO-00731 Pos 2)',
     'USD', 14_222.07,
     'KIT OPTICAL SWITCH (44110148) + MOTOR STOP SWITCH O (44107515) + SEAL KIT CSK S85 (44105627). '
     'RELAY 3 PHASE deliberately excluded.'),
    # Pemberton + Justervesenet
    ('Direct supply', 'Pemberton Seraphin can + SS316 cabinet', 'USD', 75_054.92,
     'NIST + NVLAP cert accepted; no Justervesenet (NOA/Valhall precedent).'),
    # Flowtec
    ('Direct supply', 'Flowtec water draw kit parts', 'NOK', 200_000,
     'Solenoid valve + manual instrument valves.'),
    # Free-issue
    ('Free-issue', '3× pressure transmitters', 'NOK', 180_000,
     'RFQ §1.6-8.'),
    ('Free-issue', 'B&B (DBB) valve ½″ TR2000', 'NOK', 25_000, 'RFQ §1.6-7.'),
    ('Free-issue', 'Thermowells (3× process + 1× rod)', 'NOK', 60_000,
     'TR2000 BD20X.'),
    ('Free-issue', '6Mo tubing + Hoke Gyrolok fittings', 'NOK', 80_000,
     'DEV-009 / Aker BP precedent.'),
    ('Free-issue', 'BFOU halogen-free cables', 'NOK', 50_000, 'DEV-011.'),
    ('Free-issue', 'SS316 cable glands per E-001 / TR3023', 'NOK', 15_000,
     'DEV-011.'),
    ('Free-issue', '6Mo Parker plugs (vent valves)', 'NOK', 5_000,
     'NOA punch lesson C-EI-01.'),
    ('Free-issue', 'SS316 cable tray edge protection', 'NOK', 5_000,
     'C-EI-05.'),
    ('Free-issue', '100 mm insulation material', 'NOK', 25_000, 'DEV-015.'),
    # Sub-contracts + supervision
    ('Sub-contract', 'NORSOK M-501 6C paint shop (Norway)', 'NOK', 130_000,
     'DEV-004.'),
    ('Sub-contract', 'Honeywell supervisor onshore (~120 h @ $140 + KPI)', 'USD',
     16_800, 'Quote Pos 4 hourly.'),
    ('Sub-contract', 'Honeywell supervisor offshore (~240 h @ $180 + KPI + mob)',
     'USD', 43_200, 'Quote Pos 4 hourly + mob.'),
    ('Sub-contract', 'Intertek prover expert (276 h)', 'USD', 38_640,
     'Quote Pos 4.'),
    ('Sub-contract', 'Blu Electro offshore cabling', 'NOK', 200_000,
     'Quote Pos 4.'),
    # Logistics
    ('Service', 'Air freight Honeywell → Sifab', 'USD', 58_798,
     'Quote Pos 3. Sea freight option.'),
    ('Service', 'Håkull AS onshore freight Sifab → quayside', 'NOK', 50_000,
     'Local NOR.'),
    ('Service', 'Travel — Tempe FAT', 'USD', 30_104.74, 'Quote Pos 6.'),
    # Internal Sifab labour estimate
    ('Internal', 'Sifab project management + documentation', 'NOK', 400_000,
     'Quote Pos 7 budget — Sifab internal cost.'),
    ('Internal', 'Sifab engineering hours (pre-FAT inspection AZ + GA review + SDL + procedures)',
     'NOK', 250_000, 'Lesson C-QM-01: budget travel.'),
    # Banking
    ('Banking', 'Bank guarantee fee (1 yr)', 'USD', 3_000,
     '~0.5–1%/yr on USD 482k.'),
    ('Banking', 'Credit facility setup', 'NOK', 2_000, 'Lesson C-FIN-03.'),
    # Contingency
    ('Contingency', 'VOR contingency 5% of PO', 'USD', 80_413,
     'Lesson C-FIN-01: build VOR contingency.'),
]


def build():
    wb = Workbook()

    # =====================================================================
    # SUMMARY
    # =====================================================================
    ws = wb.active
    ws.title = 'Summary'
    ws['A1'] = 'SP-01415 Snorre A SVP085 — Financial Overview (internal)'
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = TITLE_FILL
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 28

    meta = [
        ('Project', 'SP-01415 Snorre A SVP085'),
        ('Customer Project', 'GM-5341 Snorre A Oil Metering'),
        ('Customer PO', '4500998501 — USD 1,608,262.36'),
        ('Owner agent', 'commercial-finance'),
        ('Last updated', '2026-05-06'),
        ('FX assumption', f'1 USD = {USD_NOK} NOK (commercial-finance to confirm + hedge if exposure persists)'),
    ]
    for i, (k, v) in enumerate(meta):
        r = 2 + i
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
        ws.cell(row=r, column=1).alignment = WRAP
        ws.cell(row=r, column=2).alignment = WRAP

    summary_row = 2 + len(meta) + 1
    ws.cell(row=summary_row, column=1, value='Headline P&L (all in USD equivalent)').font = SECTION_FONT
    ws.cell(row=summary_row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=6)

    cost_usd = sum(c for cat, _, cur, c, _ in COSTS if cur == 'USD')
    cost_nok = sum(c for cat, _, cur, c, _ in COSTS if cur == 'NOK')
    cost_total_usd = cost_usd + (cost_nok / USD_NOK)
    margin_usd = PO_VALUE_USD - cost_total_usd
    margin_pct = margin_usd / PO_VALUE_USD * 100

    rows = [
        ('Revenue (Guidant PO)', PO_VALUE_USD, ''),
        ('Direct cost (USD)', cost_usd, ''),
        ('Direct cost (NOK)', cost_nok, f'≈ USD {cost_nok / USD_NOK:,.2f} @ {USD_NOK}'),
        ('Total cost (USD equivalent)', cost_total_usd, ''),
        ('Gross margin (USD)', margin_usd, f'{margin_pct:.1f}%'),
        ('Gross margin %', margin_pct, ''),
    ]
    for i, (label, value, note) in enumerate(rows):
        r = summary_row + 1 + i
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=r, column=2, value=value)
        if isinstance(value, (int, float)):
            c.number_format = '#,##0.00' if 'margin %' not in label.lower() else '0.00'
        ws.cell(row=r, column=3, value=note)
        for col in (1, 2, 3):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).alignment = WRAP

    note_row = summary_row + 1 + len(rows) + 2
    ws.cell(row=note_row, column=1, value='Margin notes').font = SECTION_FONT
    ws.cell(row=note_row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)

    margin_notes = [
        'Quote-stage margin (R1 Honeywell): ~USD 1,018k (~63%). R2 quote saved USD 8,577 → ~USD 1,027k.',
        'Free-issue items + Norwegian-side scope (paint, cabling, freight, internal labour) are the largest variable items.',
        'Pos 4 hourly work (HE supervisor + Intertek + Blu Electro) is the single biggest schedule + cost risk; budget vs actual to be tracked monthly.',
        'VOR contingency 5% (~USD 80k) reserved against unknown deviations; Lesson C-FIN-01.',
        f'Currency exposure: USD revenue, mixed cost (~70% USD / 30% NOK). Hedge recommended if USD/NOK moves > 5% from quote rate {USD_NOK}.',
        'Bank guarantee fee + credit facility cost included; lesson C-FIN-03 — credit facility before any supplier down-payment.',
    ]
    for i, txt in enumerate(margin_notes):
        ws.cell(row=note_row + 1 + i, column=1, value=txt).alignment = WRAP
        ws.merge_cells(
            start_row=note_row + 1 + i, start_column=1,
            end_row=note_row + 1 + i, end_column=6,
        )

    for w_idx, w in enumerate([35, 22, 35, 12, 12, 12], start=1):
        ws.column_dimensions[get_column_letter(w_idx)].width = w

    # =====================================================================
    # REVENUE
    # =====================================================================
    rev_ws = wb.create_sheet('Revenue')
    rev_ws['A1'] = 'Revenue — Guidant PO 4500998501 milestones'
    rev_ws['A1'].font = TITLE_FONT
    rev_ws['A1'].fill = TITLE_FILL
    rev_ws.merge_cells('A1:F1')

    headers = ['Milestone', 'Trigger', '%', 'Amount (USD)', 'Expected date', 'Status']
    for c_idx, h in enumerate(headers, start=1):
        cell = rev_ws.cell(row=2, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    rev_ws.row_dimensions[2].height = 30

    for i, (m, trigger, pct, dt, status, note) in enumerate(REVENUE, start=3):
        rev_ws.cell(row=i, column=1, value=m).font = Font(bold=True)
        rev_ws.cell(row=i, column=2, value=trigger)
        rev_ws.cell(row=i, column=3, value=pct).number_format = '0.0%'
        c = rev_ws.cell(row=i, column=4, value=PO_VALUE_USD * pct)
        c.number_format = '#,##0.00'
        rev_ws.cell(row=i, column=5, value=dt)
        rev_ws.cell(row=i, column=6, value=status)
        for col in range(1, 7):
            rev_ws.cell(row=i, column=col).border = BORDER
            rev_ws.cell(row=i, column=col).alignment = WRAP

    total_row = 3 + len(REVENUE)
    rev_ws.cell(row=total_row, column=1, value='Total').font = Font(bold=True)
    c = rev_ws.cell(row=total_row, column=4, value=PO_VALUE_USD)
    c.font = Font(bold=True)
    c.number_format = '#,##0.00'
    c.fill = POS_FILL

    for w_idx, w in enumerate([10, 50, 8, 16, 14, 12], start=1):
        rev_ws.column_dimensions[get_column_letter(w_idx)].width = w

    # =====================================================================
    # COSTS
    # =====================================================================
    cost_ws = wb.create_sheet('Costs')
    cost_ws['A1'] = 'Direct costs (in original currency)'
    cost_ws['A1'].font = TITLE_FONT
    cost_ws['A1'].fill = TITLE_FILL
    cost_ws.merge_cells('A1:F1')

    headers = ['Category', 'Item', 'Currency', 'Amount', 'Amount (USD eq.)', 'Notes']
    for c_idx, h in enumerate(headers, start=1):
        cell = cost_ws.cell(row=2, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER

    for i, (cat, item, cur, amt, note) in enumerate(COSTS, start=3):
        cost_ws.cell(row=i, column=1, value=cat)
        cost_ws.cell(row=i, column=2, value=item)
        cost_ws.cell(row=i, column=3, value=cur)
        c = cost_ws.cell(row=i, column=4, value=amt)
        c.number_format = '#,##0.00'
        usd_eq = amt if cur == 'USD' else amt / USD_NOK
        c2 = cost_ws.cell(row=i, column=5, value=usd_eq)
        c2.number_format = '#,##0.00'
        cost_ws.cell(row=i, column=6, value=note)
        for col in range(1, 7):
            cost_ws.cell(row=i, column=col).border = BORDER
            cost_ws.cell(row=i, column=col).alignment = WRAP

    total_row = 3 + len(COSTS)
    cost_ws.cell(row=total_row, column=2, value='TOTAL').font = Font(bold=True)
    c = cost_ws.cell(row=total_row, column=5, value=cost_total_usd)
    c.font = Font(bold=True)
    c.number_format = '#,##0.00'
    c.fill = NEG_FILL

    for w_idx, w in enumerate([14, 50, 10, 14, 16, 40], start=1):
        cost_ws.column_dimensions[get_column_letter(w_idx)].width = w

    # =====================================================================
    # CASHFLOW
    # =====================================================================
    cf_ws = wb.create_sheet('Cashflow')
    cf_ws['A1'] = 'Cashflow timeline (USD equivalent)'
    cf_ws['A1'].font = TITLE_FONT
    cf_ws['A1'].fill = TITLE_FILL
    cf_ws.merge_cells('A1:E1')

    headers = ['Month', 'In (revenue)', 'Out (cost)', 'Net month', 'Cumulative']
    for c_idx, h in enumerate(headers, start=1):
        cell = cf_ws.cell(row=2, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER

    timeline = [
        ('2026-05', PO_VALUE_USD * 0.30, 175_446 + 3_000,
         'M1 invoice + Honeywell 30% down + BG fee'),
        ('2026-06', 0, 50_000, 'Free-issue procurement starts (transmitters, B&B, RTDs)'),
        ('2026-07', 0, 30_000, 'Free-issue + thermowells + cables'),
        ('2026-08', 0, 50_000, 'Free-issue kit ships to Honeywell'),
        ('2026-09', 0, 25_000, 'Pemberton Seraphin + Flowtec PO'),
        ('2026-10', 0, 0, ''),
        ('2026-11', 0, 0, ''),
        ('2026-12', PO_VALUE_USD * 0.50, 350_891 + 30_000, 'M2 invoice + Honeywell 60% milestone + Tempe travel'),
        ('2027-01', 0, 16_800 + 13_000, 'Onshore disassembly supervision + paint shop'),
        ('2027-02', 0, 58_798, 'Air freight Honeywell → Sifab'),
        ('2027-03', 0, 5_000, 'Onshore transport to quayside'),
        ('2027-04', PO_VALUE_USD * 0.10, 43_200 + 38_640 + 18_780,
         'M3 invoice + offshore HE + Intertek + Blu Electro'),
        ('2027-05', 0, 0, ''),
        ('2027-08', PO_VALUE_USD * 0.10, 58_482, 'M4 invoice + Honeywell 10% closeout'),
    ]
    cumulative = 0.0
    for i, (m, inflow, outflow, note) in enumerate(timeline, start=3):
        net = inflow - outflow
        cumulative += net
        cf_ws.cell(row=i, column=1, value=m).font = Font(bold=True)
        cf_ws.cell(row=i, column=2, value=inflow).number_format = '#,##0.00'
        cf_ws.cell(row=i, column=3, value=outflow).number_format = '#,##0.00'
        c = cf_ws.cell(row=i, column=4, value=net)
        c.number_format = '#,##0.00'
        c.fill = POS_FILL if net >= 0 else NEG_FILL
        c2 = cf_ws.cell(row=i, column=5, value=cumulative)
        c2.number_format = '#,##0.00'
        c2.fill = POS_FILL if cumulative >= 0 else NEG_FILL
        cf_ws.cell(row=i, column=6, value=note)
        for col in range(1, 7):
            cf_ws.cell(row=i, column=col).border = BORDER
            cf_ws.cell(row=i, column=col).alignment = WRAP

    for w_idx, w in enumerate([10, 14, 14, 14, 16, 50], start=1):
        cf_ws.column_dimensions[get_column_letter(w_idx)].width = w

    # =====================================================================
    # FX
    # =====================================================================
    fx_ws = wb.create_sheet('FX')
    fx_ws['A1'] = 'USD/NOK exposure'
    fx_ws['A1'].font = TITLE_FONT
    fx_ws['A1'].fill = TITLE_FILL
    fx_ws.merge_cells('A1:D1')

    headers = ['Currency', 'Inflow', 'Outflow', 'Net']
    for c_idx, h in enumerate(headers, start=1):
        cell = fx_ws.cell(row=2, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER

    usd_in = PO_VALUE_USD
    usd_out = sum(c for cat, _, cur, c, _ in COSTS if cur == 'USD')
    nok_in = 0
    nok_out = sum(c for cat, _, cur, c, _ in COSTS if cur == 'NOK')

    for i, (cur, inflow, outflow) in enumerate(
        [('USD', usd_in, usd_out), ('NOK', nok_in, nok_out)], start=3):
        fx_ws.cell(row=i, column=1, value=cur).font = Font(bold=True)
        fx_ws.cell(row=i, column=2, value=inflow).number_format = '#,##0.00'
        fx_ws.cell(row=i, column=3, value=outflow).number_format = '#,##0.00'
        net = inflow - outflow
        c = fx_ws.cell(row=i, column=4, value=net)
        c.number_format = '#,##0.00'
        c.fill = POS_FILL if net >= 0 else NEG_FILL
        for col in range(1, 5):
            fx_ws.cell(row=i, column=col).border = BORDER

    fx_note = (
        f'\nCurrent assumption 1 USD = {USD_NOK} NOK. \n'
        f'Net USD position: USD {usd_in - usd_out:,.2f} (long USD).\n'
        f'Net NOK position: NOK {nok_in - nok_out:,.2f} (short NOK — cost only).\n'
        'Recommendation: hedge USD revenue at PO acceptance if rate moves > 5% from quote level. \n'
        'Lesson C-FIN-06: US tariff risk on Honeywell pricing — R2 quote already absorbed.'
    )
    fx_ws['A6'] = fx_note
    fx_ws['A6'].alignment = WRAP
    fx_ws.merge_cells('A6:D14')

    for w_idx, w in enumerate([10, 16, 16, 16], start=1):
        fx_ws.column_dimensions[get_column_letter(w_idx)].width = w

    wb.save(OUTPUT)
    print(f'Saved: {OUTPUT}')
    print(f'Revenue: USD {PO_VALUE_USD:,.2f}')
    print(f'Cost: USD {cost_total_usd:,.2f} (USD direct {usd_out:,.2f} + NOK {nok_out:,.2f} @ {USD_NOK})')
    print(f'Margin: USD {margin_usd:,.2f} ({margin_pct:.1f}%)')


if __name__ == '__main__':
    build()
