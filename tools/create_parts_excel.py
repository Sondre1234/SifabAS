import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

# -- Sheet 1: Parts by Supplier --
ws = wb.active
ws.title = 'Parts by Supplier'

header_font = Font(bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
title_font = Font(bold=True, size=14)
subtitle_font = Font(bold=True, size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap = Alignment(wrap_text=True, vertical='top')

# Title
ws.merge_cells('A1:F1')
ws['A1'] = 'Free-Issued Parts to Honeywell/TruStop Factory - NOA & Valhall Fenris Provers'
ws['A1'].font = title_font

ws.merge_cells('A2:F2')
ws['A2'] = 'Reference Projects: SP-00525 (Hugin A / SVP050) & SP-00577 (Valhall Fenris / SVP015)'
ws['A2'].font = subtitle_font

ws.merge_cells('A3:F3')
ws['A3'] = 'Prepared by Sifab AS - For reference on Snorre A SVP085 project (GM-8501-1447)'

# Headers
headers = ['#', 'Supplier', 'Parts Supplied', 'Specification', 'Qty per Prover', 'Notes']
for col, h in enumerate(headers, 1):
    c = ws.cell(row=5, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    c.border = thin_border

# Data
data = [
    [1, 'Draka', 'BFOU cables - optical switches', '#20110634 (1 triple), BFOU(i) S3/S7/S103 250V', '~50m', 'For 3 optical switches on switch bar'],
    [2, 'Draka', 'BFOU cables - motor stop switch', '#20110623 (1 pair), BFOU(i) S3/S7/S103 250V', '~14m', 'For motor stop limit switch'],
    [3, 'Draka', 'BFOU cables - temperature elements', '#20110623 (1 pair) + #20110626 (2 pair)', '~20m each', 'JB to transmitter + transmitter to TE'],
    [4, 'R. Stahl / Tranberg', 'Ex i terminal box (JB)', 'Type 8150/2-0360-0360-150-3311, SS 316L, IP66, Ex ia/ib IIC T5 Gb', '1', 'Assembled by Tranberg in Norway'],
    [5, 'Phoenix', 'Terminal blocks + jumpers', 'UT 2.5 BU terminal blocks, FPS jumpers', '37 blocks/JB', 'Installed inside JBs'],
    [6, 'CMP', 'Cable glands, stopping plugs, drain plugs', 'A2F100 M32/M20, 757 M32/M20, 781E M20 - nickel-plated brass', '~11 per JB', 'Fitted in JBs'],
    [7, 'B.F.E. / Bonney Forge (Italy)', '1/2" Double Block & Bleed valve', 'DN15, Duplex SS F51, NACE MR0175, extended handle for 100mm insulation', '1', 'CL 150 (Valhall) / CL 300 (Hugin A). Via MRC Global Norway'],
    [8, 'Pemberton Fabricators (NJ, USA)', 'Seraphin calibration cans', '151.4L (Hugin A) / 75.7L (Valhall). SS 316, NIST traceable', '1', 'Calibrated by Pennsylvania Standards Lab (NVLAP 200869-0)'],
    [9, 'Metallteknikk AS (Sandnes/Bryne)', 'SS 316L transport enclosures for Seraphin cans', 'Art. 143251 (1700x1000x1200mm, 235kg) / Art. 143252 (1620x850x1200mm, 208kg)', '1', 'Designed by Sifab (Idar Eriksen). EMKA locks, forklift-liftable'],
    [10, 'Lucifer (via Flowtec AS)', 'Solenoid valves for water draw kit', 'Model 201LG4SVG7, coil 492190C2, SS 316, 24VDC, 1/2" BSP, IP66, Ex eb mb IIC T3/T4', '1', 'Direct pilot operated, 2-port, spring return'],
    [11, 'Hoke', 'Tubing fittings for water draw kit', '1" 6Mo tubing and fittings (upgraded from 1/2" via VOR 003)', 'Kit', '1" fittings 3x cost of 1/2". Larger bending radius required elbows'],
    [12, 'Lanne Elektriske (Aker BP frame agr.)', 'Motor - ATB brand', '1 HP, 690VAC 60Hz 3ph (Valhall) / 690VAC 50Hz 3ph (Hugin A)', '1', 'Non-standard motors per Aker BP frame agreement. Caused delays'],
    [13, 'Sifab AS (fabricated)', 'Thermowells - Duplex SS', 'Per HW drawings. 1/2" NPTF, bore 6.5mm', '3', 'Extra TW on switch bar for annual calibration'],
    [14, 'Sifab AS (fabricated)', 'DBB valve adapter', '1/2" ANSI 600 x ANSI 150/300 (B16.5)', '1', 'Adapts prover 600# connection to DBB valve class'],
    [15, 'Sifab AS (fabricated)', 'Pipe stand for temperature transmitters', 'Custom bracket for mounting outside cabinet (insulation)', '1', 'Added via VOR 003 - insulation required external mounting'],
    [16, 'Sifab AS (fabricated)', 'Customer enclosure', 'SS 316', '1', 'Junction box mounting enclosure'],
    [17, 'Sifab AS (purchased)', 'Bolts and nuts', 'Stud bolts + hex nuts per PCS AD20/BD20. Sizes: 1/2"-13, 5/8"-11', '36-60', 'All with MTR certificates'],
    [18, 'Sifab AS (purchased)', 'Spiral wound gaskets', '1/2" 150#/300#/600# SS 316', '6 (spares)', 'Sent as extra spares'],
    [19, 'Sifab AS (purchased)', '6Mo Parker plugs', 'NPT plugs in 6Mo material', 'As needed', 'Replaced standard vent valves. Added via VOR 004'],
    [20, 'Sifab AS (purchased)', 'Cable glands for instruments', 'M25/M20 with 90-deg elbows + straight, M20 adapters', '~12', 'For temp transmitters, switches, JB entries'],
    [21, 'FMC (free-issued direct)', '4-wire RTDs', 'Similar to HW P/N 44102764/44102765/44110604', '3', 'Sent direct from FMC manufacturer, not via Sifab'],
    [22, 'FMC (free-issued direct)', 'Temperature transmitters', 'Prover TT, Switch bar TT, Extra TT', '3', 'Sent direct from FMC manufacturer'],
    [23, 'FMC (free-issued direct)', 'Pressure transmitters', 'Not in Honeywell scope - FMC installs', '1', 'FMC responsibility'],
]

for i, row in enumerate(data):
    for j, val in enumerate(row):
        c = ws.cell(row=6+i, column=j+1, value=val)
        c.border = thin_border
        c.alignment = wrap

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 50
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 45


# -- Sheet 2: Snorre A Procurement Checklist --
ws2 = wb.create_sheet('Snorre A Checklist')

ws2.merge_cells('A1:F1')
ws2['A1'] = 'Snorre A SVP085 - Procurement Checklist (Based on NOA/Valhall Experience)'
ws2['A1'].font = title_font

ws2.merge_cells('A2:F2')
ws2['A2'] = 'Items Sifab will likely need to source and free-issue to Honeywell/TruStop'

headers2 = ['#', 'Item', 'Expected Spec (Snorre A)', 'Supplier (NOA/Valhall)', 'Action Required', 'Status']
for col, h in enumerate(headers2, 1):
    c = ws2.cell(row=4, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    c.border = thin_border

checklist = [
    [1, 'BFOU cables (switches + temp + motor)', 'BFOU type per TR3023/E-001. Confirm lengths with HW GA', 'Draka', 'Get cable schedule from Honeywell after GA drawing', 'TBD'],
    [2, 'Ex terminal box (JB)', 'Ex ia/ib IIC, SS 316L, IP66. Size TBD based on cable count', 'R. Stahl / Tranberg', 'Confirm JB size after wiring diagram', 'TBD'],
    [3, 'Terminal blocks + jumpers', 'Phoenix UT 2.5 BU + FPS jumpers per termination diagram', 'Phoenix', 'Order after termination diagram approved', 'TBD'],
    [4, 'Cable glands', 'M25/M20/M32 nickel-plated brass or SS 316 per E-001', 'CMP', 'Confirm sizes with JB and instrument entries', 'TBD'],
    [5, 'DBB valve', '1/2" CL 600 (BD20X), Duplex SS F51, NACE, extended for insulation', 'B.F.E. / Bonney Forge via MRC Global', 'LONG LEAD - order early. Confirm class with Guidant', 'TBD'],
    [6, 'DBB valve adapter', '1/2" ANSI 600 adapter if needed', 'Sifab (fabricate)', 'Confirm if needed based on prover connection', 'TBD'],
    [7, 'Seraphin can', 'Size TBD per SVP085 volume - likely larger than 151.4L', 'Pemberton Fabricators (NJ)', 'Get size from Honeywell. Justervesenet cert required', 'TBD'],
    [8, 'Seraphin transport enclosure', 'SS 316L, sized for Seraphin can, forklift-liftable', 'Metallteknikk AS (Sandnes)', 'Design after can size confirmed', 'TBD'],
    [9, 'Solenoid valve (water draw kit)', 'SS 316, 24VDC, Ex rated. Confirm orifice size early!', 'Lucifer via Flowtec AS', 'Specify orifice in initial design - avoid VOR', 'TBD'],
    [10, 'Water draw kit tubing + fittings', '1" 6Mo (use 1" from start - lesson learned from VOR)', 'Hoke', 'Specify 1" from start - the upgrade was biggest VOR item', 'TBD'],
    [11, 'Motor', '230VAC 3ph 60Hz (or 50Hz - TQ-001 pending)', 'Honeywell (standard)', 'If standard HW motor, no free-issue needed. Confirm freq', 'TBD'],
    [12, 'Thermowells - Duplex SS', 'Per BD20X MDS. 1/2" NPTF, bore 6.5mm. 3x process + 1x rod', 'Sifab (fabricate)', 'Get TW drawings from HW. Fabricate in Norway, ship to AZ', 'TBD'],
    [13, 'Bolts, nuts, gaskets', 'Per BD20X piping class (CL 600). Studs + hex nuts with MTRs', 'Sifab (purchase)', 'Get bolt list from HW after GA approved', 'TBD'],
    [14, 'Pipe stand for temp transmitters', 'Custom bracket for external mounting (insulation)', 'Sifab (fabricate)', 'Include in initial design - was VOR on previous projects', 'TBD'],
    [15, '6Mo Parker plugs', 'NPT plugs in 6Mo per TR2000 BD20X', 'Sifab (purchase)', 'Include from start - was VOR 004 on previous projects', 'TBD'],
    [16, 'Temperature elements (RTDs)', '4-wire RTDs - Guidant free-issue per RFQ', 'Guidant (free-issue)', 'Confirm with Guidant - vendor installs TE but does not supply', 'TBD'],
    [17, 'Temperature transmitters', 'Guidant free-issue per RFQ', 'Guidant (free-issue)', 'Confirm with Guidant', 'TBD'],
]

for i, row in enumerate(checklist):
    for j, val in enumerate(row):
        c = ws2.cell(row=5+i, column=j+1, value=val)
        c.border = thin_border
        c.alignment = wrap

ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 35
ws2.column_dimensions['C'].width = 50
ws2.column_dimensions['D'].width = 30
ws2.column_dimensions['E'].width = 45
ws2.column_dimensions['F'].width = 12

filepath = 'projects/snorre-a-compact-prover/engineering/Free_Issued_Parts_NOA_Valhall_Reference.xlsx'
wb.save(filepath)
print(f'Saved: {filepath}')
